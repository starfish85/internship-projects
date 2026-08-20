#!/usr/bin/env python3
"""Write hotel product workbook: sheet1 = product fields, other sheets = calendar by month."""
from __future__ import annotations

import argparse
import json
import sys
from collections import defaultdict
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font
except ImportError:
    print("openpyxl required: pip3 install openpyxl --user", file=sys.stderr)
    sys.exit(1)


def flatten_product(product: dict) -> list[tuple[str, object]]:
    """Prefer stable column order for main product fields."""
    preferred = [
        "hotelCode",
        "productCode",
        "hotelCode",
        "productCode",
        "hotelbedsCode",
        "nameZh",
        "nameEn",
        "descriptionZh",
        "descriptionEn",
        "addressZh",
        "addressEn",
        "cityZh",
        "cityEn",
        "contentStatus",
        "contentMissing",
        "hasClientIntro",
        "languageFallback",
        "categoryCode",
        "categorySimpleCode",
        "categoryZh",
        "categoryEn",
        "categoryGroupCode",
        "categoryGroupZh",
        "categoryGroupEn",
        "stars",
        "halfStar",
        "chainCode",
        "chainZh",
        "chainEn",
        "accommodationTypeCode",
        "accommodationTypeZh",
        "accommodationTypeEn",
        "countryId",
        "countryCode",
        "countryIsoCode",
        "countryZh",
        "countryEn",
        "destinationId",
        "destinationCode",
        "destinationZh",
        "destinationEn",
        "zoneId",
        "zoneCode",
        "zoneZh",
        "zoneEn",
        "postalCode",
        "latitude",
        "longitude",
        "coordinateFix",
        "phone",
        "phoneHotel",
        "phoneBooking",
        "fax",
        "email",
        "website",
        "amenitiesZh",
        "amenitiesEn",
        "amenityCodes",
        "facilities",
        "facilitiesZh",
        "facilitiesEn",
        "segmentsZh",
        "segmentsEn",
        "segmentCodes",
        "exclusiveDeal",
        "luxuryCollection",
        "sustainable",
        "vacationRental",
        "top",
        "mostPopular",
        "deposit",
        "urlImage",
        "urlImageBigger",
        "urlImageRel",
        "imageCount",
        "imagePaths",
        "imageLocalStatus",
        "detailLevel",
        "source",
        "capturedAt",
        "contentUpdatedAt",
        "notes",
    ]
    rows: list[tuple[str, object]] = []
    seen = set()
    for k in preferred:
        if k in product:
            v = product[k]
            if isinstance(v, (list, dict)):
                v = json.dumps(v, ensure_ascii=False)
            rows.append((k, v))
            seen.add(k)
    for k, v in sorted(product.items()):
        if k in seen:
            continue
        if isinstance(v, (list, dict)):
            v = json.dumps(v, ensure_ascii=False)
        rows.append((k, v))
    return rows


def write_workbook(product_path: Path, calendar_path: Path | None, out_path: Path) -> dict:
    product = json.loads(product_path.read_text(encoding="utf-8"))
    calendar_rows = []
    if calendar_path and calendar_path.exists():
        raw = json.loads(calendar_path.read_text(encoding="utf-8"))
        if isinstance(raw, list):
            calendar_rows = raw
        elif isinstance(raw, dict):
            calendar_rows = raw.get("rates") or raw.get("days") or raw.get("items") or []

    wb = Workbook()
    ws0 = wb.active
    ws0.title = "酒店产品信息"
    ws0["A1"] = "字段"
    ws0["B1"] = "值"
    ws0["A1"].font = Font(bold=True)
    ws0["B1"].font = Font(bold=True)
    for i, (k, v) in enumerate(flatten_product(product), start=2):
        ws0[f"A{i}"] = k
        ws0[f"B{i}"] = "" if v is None else v

    # Group calendar by YYYY-MM
    by_month: dict[str, list[dict]] = defaultdict(list)
    for row in calendar_rows:
        date = row.get("date") or row.get("checkIn") or row.get("day") or ""
        month = str(date)[:7] if date else "unknown"
        by_month[month].append(row)

    if not by_month:
        ws = wb.create_sheet("日历价_暂无")
        ws["A1"] = "说明"
        ws["B1"] = "尚无日历价数据（探针未命中或未释放库存）"

    for month in sorted(by_month.keys()):
        title = f"日历价_{month}".replace("/", "-")[:31]
        ws = wb.create_sheet(title)
        rows = by_month[month]
        # union keys
        keys = []
        seen = set()
        preferred_cal = [
            "date",
            "checkIn",
            "checkOut",
            "nights",
            "board",
            "boardName",
            "roomType",
            "roomName",
            "rateKey",
            "price",
            "currency",
            "tax",
            "status",
            "allotment",
            "refundable",
            "raw",
        ]
        for k in preferred_cal:
            if any(k in r for r in rows):
                keys.append(k)
                seen.add(k)
        for r in rows:
            for k in r.keys():
                if k not in seen:
                    keys.append(k)
                    seen.add(k)
        for c, k in enumerate(keys, start=1):
            cell = ws.cell(1, c, k)
            cell.font = Font(bold=True)
        for ri, r in enumerate(rows, start=2):
            for ci, k in enumerate(keys, start=1):
                v = r.get(k)
                if isinstance(v, (list, dict)):
                    v = json.dumps(v, ensure_ascii=False)
                ws.cell(ri, ci, v)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return {
        "out": str(out_path),
        "productFields": len(product),
        "calendarRows": len(calendar_rows),
        "months": sorted(by_month.keys()),
    }


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--product", required=True, help="product JSON path")
    ap.add_argument("--calendar", default="", help="calendar JSON path (list or {rates:[]})")
    ap.add_argument("--out", required=True, help="output xlsx path")
    args = ap.parse_args()
    result = write_workbook(
        Path(args.product),
        Path(args.calendar) if args.calendar else None,
        Path(args.out),
    )
    print(json.dumps(result, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
