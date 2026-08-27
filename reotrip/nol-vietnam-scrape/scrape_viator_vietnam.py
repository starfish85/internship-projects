#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Viator 越南单门票 / 纯接送采集与导出。入口见前端搜索与分类页。"""

from __future__ import annotations

import json
import re
import time
from pathlib import Path
from typing import Any, Dict, List, Optional
from urllib.parse import quote_plus

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

ROOT = Path(__file__).resolve().parent
DATA = ROOT / "data" / "viator_listings.json"
OUT = ROOT / "Viator_越南单门票纯接送.xlsx"
SITE = "https://www.viator.com"

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(name="宋体", bold=True, color="FFFFFF", size=11)
DATA_FONT = Font(name="宋体", size=11)
HEADER_AL = Alignment(wrap_text=True, vertical="center")
DATA_AL = Alignment(wrap_text=True, vertical="top")
THIN = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)

PRODUCT_HEADERS = [
    "分类", "产品ID", "链接", "城市（英）", "城市（中）", "标题（英）", "标题（中）",
    "副标题（英）", "副标题（中）", "格式（英）", "格式（中）", "成团类型（英）", "成团类型（中）",
    "出行方式（英）", "出行方式（中）", "时长(小时)", "最多人数", "可带儿童", "语言（英）", "语言（中）",
    "价格", "货币", "计价单位（英）", "计价单位（中）", "价格展示（英）", "已下单人数", "前端口径",
    "评论数", "评分", "评分20", "热度", "即时确认", "状态", "集合点（英）", "集合点（中）",
    "简介（英）", "简介（中）", "描述（英）", "描述（中）", "套餐数", "封面图", "组织者",
    "组织者订单数", "组织者接待人数", "组织者已付款人数", "组织者评分", "标签（英）", "可订日期数", "抓取时间",
]
OPTION_HEADERS = [
    "分类", "产品ID", "产品标题（英）", "产品标题（中）", "套餐ID", "套餐名（英）", "套餐名（中）",
    "套餐说明（英）", "套餐说明（中）", "计价类型（英）", "计价类型（中）", "价格", "货币",
    "展示（英）", "单位（英）", "最少数量", "最多数量", "提前截止(小时)",
]
TRAFFIC_HEADERS = [
    "分类", "产品ID", "标题（英）", "标题（中）", "城市（中）", "已下单人数",
    "前端文案（英）", "前端文案（中）", "评论数", "评分", "评分20", "热度", "组织者",
    "组织者累计订单", "组织者累计接待", "组织者已付款人数", "组织者评分", "组织者评论数", "链接",
]
PRODUCT_WIDTHS = {
    "A": 12, "B": 16, "C": 42, "D": 22, "E": 13, "F": 42, "G": 42, "H": 36, "I": 36,
    "J": 16, "K": 14, "L": 14, "M": 20, "N": 14, "O": 14, "P": 12, "Q": 12, "R": 12,
    "S": 14, "T": 14, "U": 10, "V": 10, "W": 16, "X": 14, "Y": 18, "Z": 14,
    "AA": 28, "AB": 10, "AC": 10, "AD": 10, "AE": 10, "AF": 12, "AG": 12,
    "AH": 28, "AI": 28, "AJ": 40, "AK": 40, "AL": 40, "AM": 40, "AN": 10, "AO": 36,
    "AP": 20, "AQ": 12, "AR": 12, "AS": 12, "AT": 12, "AU": 28, "AV": 12, "AW": 20,
}
OPTION_WIDTHS = {
    "A": 12, "B": 16, "C": 42, "D": 42, "E": 16, "F": 36, "G": 36, "H": 36, "I": 36,
    "J": 18, "K": 14, "L": 10, "M": 10, "N": 18, "O": 12, "P": 10, "Q": 10, "R": 14,
}
TRAFFIC_WIDTHS = {
    "A": 12, "B": 16, "C": 42, "D": 42, "E": 13, "F": 14, "G": 14, "H": 22, "I": 10,
    "J": 10, "K": 10, "L": 10, "M": 20, "N": 14, "O": 14, "P": 16, "Q": 12, "R": 12, "S": 42,
}

CITY_ZH = {
    "Ho Chi Minh City": "胡志明市", "Hanoi": "河内", "Da Nang": "岘港", "Hoi An": "会安",
    "Nha Trang": "芽庄", "Phu Quoc": "富国岛", "Hue": "顺化", "Da Lat": "大叻", "Sapa": "沙巴",
    "Ha Long": "下龙", "Tay Ninh": "西宁", "Quy Nhon": "归仁", "Ninh Binh": "宁平", "Vietnam": "越南",
}

PHRASES = [
    ("General Admission Ticket", "普通门票"),
    ("General Admission", "普通门票"),
    ("Toyota Vios 4 Seat", "丰田 Vios 4座"),
    ("Toyota Innova 7 Seat", "丰田 Innova 7座"),
    ("Fortuner 7 Seat", "Fortuner 7座"),
    ("4 Seater Car (Sedan)", "4座轿车"),
    ("7 Seater Car (SUV/MPV)", "7座SUV/MPV"),
    ("16 Seater Car", "16座中巴"),
    ("Private car (per group)", "包车（按车）"),
    ("All-Inclusive Admission Ticket", "一价全包门票"),
    ("Admission Ticket", "门票"),
    ("Entry Ticket", "门票"),
    ("Skip the Line", "快速通道"),
    ("Convenient Delivery Service", "便捷配送服务"),
    ("Private Transfer Between", "包车接送："),
    ("Private Transfer From", "包车接送：从"),
    ("Private Transfer from", "包车接送：从"),
    ("Private Transfer:", "包车接送："),
    ("Private Airport Transfer Service", "机场包车接送"),
    ("Private Airport Transfer", "机场包车接送"),
    ("Airport Private Arrival Transfer", "机场包车接机"),
    ("Airport Private Transfer", "机场包车接送"),
    ("Private Arrival Transfer", "包车接机"),
    ("Airport Transfer", "机场接送"),
    ("Private Transfers", "包车接送"),
    ("Private Transfer", "包车接送"),
    ("Private car transfer", "包车接送"),
    ("Private car", "包车"),
    ("Limousine Transfer Shuttle", "豪华车接送班车"),
    ("Limousine Transfer", "豪华车接送"),
    ("Luxury Limousine Vans", "豪华商务车"),
    ("by Private Car Transfers", "包车接送"),
    ("By Private Car Transfers", "包车接送"),
    ("By Private Transfer", "包车接送"),
    ("by Private Transfer", "包车接送"),
    ("by private cars", "包车"),
    ("by private car", "包车"),
    ("or vice versa", "或反向"),
    ("Vice versa", "往返"),
    ("vice versa", "往返"),
    ("One Way", "单程"),
    ("one way", "单程"),
    ("Round-Trip", "往返"),
    ("Free Hotel Pickup", "含酒店接送"),
    ("English Speaking Driver", "英语司机"),
    ("Sun World Ba Na Hills", "太阳世界巴拿山"),
    ("Ba Na Hills Alpine Coaster", "巴拿山高山滑车"),
    ("Ba Na Hills", "巴拿山"),
    ("VinWonders Nam Hoi An Theme Park", "会安南美珠乐园"),
    ("Vinwonders Nam Hoi An Theme Park", "会安南美珠乐园"),
    ("VinWonders Nam Hoi An", "会安南美珠乐园"),
    ("VinWonders Nha Trang", "芽庄美珠乐园"),
    ("Hoi An Memories Show", "会安记忆秀"),
    ("Hoi An Impression Theme Park", "会安印象主题公园"),
    ("Hoi An Memories Land", "会安记忆乐园"),
    ("Thang Long Water Puppet Theater", "升龙水上木偶剧院"),
    ("Thang Long Water Puppet Show", "升龙水上木偶演出"),
    ("Thang Long Water Puppet", "升龙水上木偶"),
    ("Golden Dragon Water Puppet Show", "金龙水上木偶演出"),
    ("Golden Dragon Water Puppet", "金龙水上木偶"),
    ("Hanoi Water Puppet Show", "河内水上木偶演出"),
    ("Hanoi Water Puppet", "河内水上木偶"),
    ("Saigon Skydeck", "西贡空中观景台"),
    ("Bitexco Financial Tower", "Bitexco金融塔"),
    ("Landmark 81 Saigon Skyview", "地标81西贡天际观景"),
    ("Landmark 81", "地标81"),
    ("Sun World Fansipan Legend", "太阳世界番西邦传奇"),
    ("Sun World Fansipan", "太阳世界番西邦"),
    ("Sun World Hon Thom Park Pass", "太阳世界Hon Thom园票"),
    ("Sun World Hon Thom", "太阳世界Hon Thom"),
    ("Sun World Ha Long", "太阳世界下龙"),
    ("Sun World Ba Den Mountain", "太阳世界黑婆山"),
    ("Commander Cave Museum", "指挥洞博物馆"),
    ("Da Nang Museum", "岘港博物馆"),
    ("Lotte World Aquarium Hanoi", "河内乐天世界水族馆"),
    ("KidZania", "KidZania职业体验城"),
    ("Dalat Wonderland", "大叻仙境乐园"),
    ("Langbiang Land Dalat", "大叻郎边乐园"),
    ("Grand World Phu Quoc Teddy Bear Museum", "富国Grand World泰迪熊博物馆"),
    ("Suoi Tien Theme Park", "仙泉主题公园"),
    ("Fito Museum Featuring Traditional Medicine", "传统医学Fito博物馆"),
    ("Fito Museum", "Fito博物馆"),
    ("CSO Gallery Admission Ticket: Coin & Stamp Museum", "CSO画廊门票：钱币邮票博物馆"),
    ("Chum Show", "Chum秀"),
    ("Life Puppets show - Đó Theatre Ticket", "生命木偶秀（Đó剧院）门票"),
    ("Alpine Coaster Experience", "高山滑车体验"),
    ("Theme Park", "主题公园"),
    ("Highland Theme Park", "高原主题公园"),
    ("Cable Car Access", "含缆车"),
    ("Cable Car", "缆车"),
    ("QRCODE Ticket", "二维码门票"),
    ("Park Pass", "园票"),
    ("Attraction & Outdoor Rides", "景点与户外游乐"),
    ("Ho Chi Minh City", "胡志明市"),
    ("Ho Chi Minh", "胡志明市"),
    ("Tan Son Nhat", "新山一"),
    ("Noi Bai", "内排"),
    ("Cam Ranh", "金兰"),
    ("Phu Cat", "符吉"),
    ("Vinpearl Phu Quoc", "富国珍珠岛"),
    ("Vinpearl Port", "珍珠港"),
    ("Tuan chau", "团洲"),
    ("Old Quarter", "老城区"),
    ("City Center", "市中心"),
    ("city center", "市中心"),
    ("Da Nang International Airport", "岘港国际机场"),
    ("Da Nang Airport", "岘港机场"),
    ("Danang Airport", "岘港机场"),
    ("Hanoi Airport", "河内机场"),
    ("Hue Airport", "顺化机场"),
    ("Phu Quoc Airport", "富国机场"),
    ("Nha Trang Airport", "芽庄机场"),
    ("Saigon Airport", "西贡机场"),
    ("SGN Airport", "新山一机场"),
    ("Airport (HAN)", "机场（HAN）"),
    ("Airport (SGN)", "机场（SGN）"),
    ("Airport (DAD)", "机场（DAD）"),
    ("(HAN)", "（HAN）"),
    ("(SGN)", "（SGN）"),
    ("(DAD)", "（DAD）"),
    ("(HUI)", "（HUI）"),
    ("(UIH)", "（UIH）"),
    ("(PQC)", "（PQC）"),
    ("(CXR)", "（CXR）"),
    ("Da Nang", "岘港"),
    ("Danang", "岘港"),
    ("Hoi An", "会安"),
    ("Hanoi", "河内"),
    ("Ha Noi", "河内"),
    ("Nha Trang", "芽庄"),
    ("Phu Quoc", "富国岛"),
    ("Hue", "顺化"),
    ("Sapa", "沙巴"),
    ("Ha Long", "下龙"),
    ("Halong", "下龙"),
    ("Ninh Binh", "宁平"),
    ("Ha Giang", "河江"),
    ("Cao Bang", "高平"),
    ("Quy Nhon", "归仁"),
    ("Tay Ninh", "西宁"),
    ("Quang Ninh", "广宁"),
    ("Vietnam", "越南"),
    ("Arrival Or Departure", "接机或送机"),
    ("Arrival or Departure", "接机或送机"),
    ("Pick up or Drop off", "接机或送机"),
    ("Pick-up", "接机"),
    ("Pickup", "接机"),
    ("Drop off", "送机"),
    ("Departure", "送机"),
    ("Arrival", "接机"),
    ("Hotel", "酒店"),
    ("Hotels", "酒店"),
    ("PRIVATECAR", "包车"),
    ("PRIVATE", "包车"),
    ("Private", "包车"),
    ("Transfer", "接送"),
    ("Transfers", "接送"),
    ("Ticket", "门票"),
    ("Tickets", "门票"),
    ("Show", "演出"),
    ("Museum", "博物馆"),
    ("Aquarium", "水族馆"),
    ("Theater", "剧院"),
    ("Theatre", "剧院"),
    ("from/to", "往返"),
    ("TO/from", "往返"),
    ("FROM/to", "往返"),
    ("to/from", "往返"),
    ("or RETURN", "或返程"),
    ("Per group", "按车"),
    ("Free Cancellation", "免费取消"),
]


FIELD_ROWS = [
    ["分类", "本地规则", "单门票：标题含 Admission/Ticket/Entry/Skip the Line，且不是一日游、jeep团、含接送的乐园套票。纯接送：Airport/Private Transfer 点对点，不含 Fast Track 通关、签证、layover 半日游。"],
    ["标题 / 简介", "Viator 搜索与分类页、产品详情页", "英文原文 + 中文对照。平台为英语站。中文为词表对照，专名可保留原文。"],
    ["链接", "www.viator.com", "有产品码写详情 URL；否则写搜索页。国家 dest=d21，胡志明 d352，河内 d351，岘港 d4680，会安 d5229。"],
    ["价格", "列表起价 from $", "货币以页面为准，多为 USD。部分分类页会切到 CAD，已按美元页回写。"],
    ["套餐", "产品页选项或列表起价", "能解析到车型（如 Vios 4 座、7 座、16 座）则分行；否则按产品起价一行。"],
    ["已下单人数 / 热度 / 组织者累计订单", "Viator 前端未统一公开", "评论数、评分来自列表/详情。本机 curl 被 DataDome 拦，本次用分类/搜索页 + 部分详情页。"],
    ["入口", "前端", "搜索 https://www.viator.com/searchResults/all?text=Vietnam+admission+ticket ；门票 /Vietnam-tours/Sightseeing-Tickets-and-Passes/d21-g8 ；主题公园 /Vietnam-tours/Theme-Park-Tickets-and-Tours/d21-g14-c50 ；交通 /Vietnam-tours/Transfers-and-Ground-Transport/d21-g15 ；机场接送 /{City}-tours/Airport-and-Ground-Transfers/d{id}-g15-c52"],
]


def translate(text: str) -> str:
    if not text:
        return ""
    out = text
    for en, zh in sorted(PHRASES, key=lambda x: len(x[0]), reverse=True):
        out = out.replace(en, zh)
    return out


def hours_of(dur: str) -> Optional[float]:
    if not dur:
        return None
    nums = [float(x) for x in re.findall(r"\d+(?:\.\d+)?", dur)]
    if not nums:
        return None
    low = dur.lower()
    if "day" in low:
        return nums[-1] * 8
    if "minute" in low and "hour" not in low:
        return round(nums[-1] / 60, 2)
    return nums[-1]


def style_sheet(ws: Worksheet, headers: List[str], widths: Dict[str, float]) -> None:
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 28
    last = get_column_letter(len(headers))
    ws.auto_filter.ref = f"A1:{last}{ws.max_row}"
    for col, h in enumerate(headers, 1):
        cell = ws.cell(1, col, h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_AL
        cell.border = THIN
        ws.column_dimensions[get_column_letter(col)].width = widths.get(get_column_letter(col), 14)
    for r in range(2, ws.max_row + 1):
        for c in range(1, len(headers) + 1):
            cell = ws.cell(r, c)
            cell.font = DATA_FONT
            cell.alignment = DATA_AL
            cell.border = THIN


def product_url(p: dict) -> str:
    u = (p.get("url") or "").strip()
    if u.startswith("http"):
        return u
    if u.startswith("/"):
        return SITE + u
    title = p.get("title") or ""
    return f"{SITE}/searchResults/all?text={quote_plus(title)}"


def product_id(p: dict) -> str:
    code = (p.get("code") or "").strip()
    if code:
        return code
    return re.sub(r"[^A-Za-z0-9]+", "-", p.get("title") or "")[:40].strip("-")


def group_labels(kind: str, p: dict) -> tuple[str, str, str, str]:
    shuttle = bool(p.get("shuttle")) or ("shuttle" in (p.get("title") or "").lower() and "private" not in (p.get("title") or "").lower())
    if kind == "纯接送":
        if shuttle:
            return "shared", "拼车/班车", "car", "车辆"
        return "private", "包车", "car", "车辆"
    return "private", "自行入园（单门票）", "foot", "自行前往"


def load_products() -> List[dict]:
    raw = json.loads(DATA.read_text(encoding="utf-8"))
    out: List[dict] = []
    for kind_key, kind in (("tickets", "单门票"), ("transfers", "纯接送")):
        for p in raw["entries"][kind_key]:
            item = dict(p)
            item["kind"] = kind
            item.setdefault("cur", "USD")
            item.setdefault("lang", "English")
            item.setdefault("reviews", 0)
            item.setdefault("cancel", False)
            item.setdefault("per_group", False if kind == "单门票" else True)
            out.append(item)
    return out


def main() -> None:
    scraped_at = time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime())
    seen = set()
    rows_p: List[List[Any]] = []
    rows_o: List[List[Any]] = []
    rows_t: List[List[Any]] = []

    for p in load_products():
        key = (p.get("code") or p["title"]).strip().lower()
        if key in seen:
            continue
        seen.add(key)
        kind = p["kind"]
        city_en = p["city"]
        city_zh = CITY_ZH.get(city_en, city_en)
        title_en = p["title"]
        title_zh = translate(title_en)
        pid = product_id(p)
        url = product_url(p)
        shuttle = bool(p.get("shuttle"))
        per_group = bool(p.get("per_group")) and not shuttle
        if kind == "纯接送" and not shuttle:
            per_group = True
        price = p.get("price")
        cur = p.get("cur") or "USD"
        unit_en = "per group" if per_group else "per person"
        unit_zh = "每车" if per_group else "每人"
        unit_sfx = "/group" if per_group else "/person"
        show = f"{cur} {price}{unit_sfx}" if price is not None else ""
        opts = p.get("options") or []
        n_opts = max(1, len(opts))
        g_en, g_zh, tr_en, tr_zh = group_labels(kind, p)
        fmt_en = "transfer" if kind == "纯接送" else "ticket"
        overview = p.get("overview") or ""
        meeting = p.get("meeting") or ""
        lang = p.get("lang") or "English"
        row = [
            kind, pid, url, city_en, city_zh, title_en, title_zh,
            overview, translate(overview),
            fmt_en, kind, g_en, g_zh, tr_en, tr_zh, hours_of(p.get("duration") or ""),
            None, "是", lang, translate(lang),
            price, cur, unit_en, unit_zh, show, None,
            "列表页起价 / 前端 Featured 排序价",
            p.get("reviews") or 0, p.get("rating"), p.get("rating"), None,
            "是" if p.get("cancel") else "否", "active",
            meeting, translate(meeting),
            overview, translate(overview),
            overview, translate(overview),
            n_opts, p.get("image") or "", p.get("supplier") or "",
            None, None, None, None, fmt_en, None, scraped_at,
        ]
        rows_p.append(row)
        rows_t.append([
            kind, pid, title_en, title_zh, city_zh, None,
            "reviews", "评论数（前端展示）", p.get("reviews") or 0, p.get("rating"), p.get("rating"),
            None, p.get("supplier") or "", None, None, None, None, p.get("reviews") or 0, url,
        ])
        if opts:
            for i, item in enumerate(opts):
                on, op, og = item[0], item[1], item[2]
                ou = "per_group_price" if og else "per_person_price"
                ouz = "按车计价" if og else "按人计价"
                osfx = "/group" if og else "/person"
                pr = op if op is not None else price
                rows_o.append([
                    kind, pid, title_en, title_zh, f"{pid}-{i + 1}", on, translate(on),
                    "", "", ou, ouz, pr, cur,
                    f"{cur} {pr}{osfx}" if pr is not None else "", osfx, 1, None, 24,
                ])
        else:
            rows_o.append([
                kind, pid, title_en, title_zh, pid, title_en, title_zh,
                overview, translate(overview),
                "per_group_price" if per_group else "per_person_price",
                "按车计价" if per_group else "按人计价", price, cur, show, unit_sfx, 1, None, 24,
            ])

    wb = Workbook()
    ws = wb.active
    ws.title = "产品对照"
    ws.append(PRODUCT_HEADERS)
    for r in rows_p:
        ws.append(r)
    style_sheet(ws, PRODUCT_HEADERS, PRODUCT_WIDTHS)

    ws = wb.create_sheet("套餐对照")
    ws.append(OPTION_HEADERS)
    for r in rows_o:
        ws.append(r)
    style_sheet(ws, OPTION_HEADERS, OPTION_WIDTHS)

    ws = wb.create_sheet("流量与订单")
    ws.append(TRAFFIC_HEADERS)
    for r in rows_t:
        ws.append(r)
    style_sheet(ws, TRAFFIC_HEADERS, TRAFFIC_WIDTHS)

    ws = wb.create_sheet("字段说明")
    ws.append(["字段", "来源", "说明"])
    for row in FIELD_ROWS:
        ws.append(row)
    style_sheet(ws, ["字段", "来源", "说明"], {"A": 28, "B": 32, "C": 80})

    n_tix = sum(1 for r in rows_p if r[0] == "单门票")
    n_tr = sum(1 for r in rows_p if r[0] == "纯接送")
    wb.save(OUT)
    print(f"wrote {OUT} tickets={n_tix} transfers={n_tr} options={len(rows_o)}")


if __name__ == "__main__":
    main()
