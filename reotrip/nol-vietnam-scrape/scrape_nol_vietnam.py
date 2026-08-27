#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""NOL(야놀자) 프론트 API에서 베트남 단입장권·순픽업 상품을 수집해 xlsx로 저장."""

from __future__ import annotations

import json
import random
import re
import ssl
import time
import urllib.error
import urllib.parse
import urllib.request
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import date, timedelta
from pathlib import Path
from threading import Lock
from typing import Any, Dict, Iterable, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from ko_zh import CITY_ZH, CATEGORY_ZH, translate_ko, translate_list

ROOT = Path(__file__).resolve().parent
DATA = ROOT / "data"
LISTING_PATH = DATA / "listings.json"
DETAIL_DIR = DATA / "details"
ITEM_DIR = DATA / "items"
XLSX_PATH = ROOT / "NOL_越南单门票纯接送.xlsx"

BASE = "https://tour.yanolja.com/tna/api/tour-api"
PRODUCT_URL = "https://tour.yanolja.com/tna/products/{id}"

TICKET_CAT = "6a831ffc-5ddf-4d61-870b-269be92fc21b"
TRANSFER_CAT = "257cf018-d16d-436f-adef-797fc9e4911f"
VN_ISO1 = "6096fa84-becb-40ae-b8f2-079288204f11"

# keyword=베트남 필터에서 확인한 도시 ID (2026-08-26)
SEED_CITIES = {
    "호치민": "f486f75f-58cc-419c-b9de-16eafe94bcf6",
    "다낭": "80397537-a64f-4a1f-a55f-411a67f9f4ef",
    "하노이": "ad4c72b3-894f-4fcc-bbff-13993e5a111a",
    "호이안": "5db37216-00a2-45de-b0e4-ef338c31e170",
    "나트랑": "3722126c-5b89-49c1-806f-414823197d3c",
    "푸꾸옥": "e985cef4-7663-40b8-9d44-1b640e0d32d8",
    "후에": "0c418fdb-8037-4008-a4ef-255730735958",
    "달랏": "871dc087-825c-41a3-882d-cb312407ee07",
    "닌빈": "c74c86ed-ea8e-4afe-a8db-e5927459fc02",
    "무이네": "010c9841-e548-4a67-9c27-d7f67d89e90b",
    "하이퐁": "8c3a2eca-d72f-421a-b7ed-8288c95f2499",
    "판티엣": "03204435-5d8c-474f-96e1-1a5aa2a4ca6f",
    "몽까이": "8eabca17-f35d-4ae9-b734-99ca1931af6f",
    "동허이": "e002b01f-4e57-4fcf-b98f-a1c16b52d684",
    "미토": "d9dd384a-385d-4c08-8d3b-9b52828f2a82",
    "사파": "be4f29ec-3b9a-4f33-beb0-d54beffced22",
}

UA = (
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
    "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
)

CTX = ssl.create_default_context()
_PRINT_LOCK = Lock()
_REQ_LOCK = Lock()
_LAST_REQ = 0.0
MIN_INTERVAL = 0.18

START_DATE = (date.today() + timedelta(days=2)).isoformat()
END_DATE = (date.today() + timedelta(days=8)).isoformat()


def log(msg: str) -> None:
    with _PRINT_LOCK:
        print(msg, flush=True)


def request_json(url: str, method: str = "GET", data: Optional[dict] = None, retries: int = 4) -> Any:
    global _LAST_REQ
    body = None if data is None else json.dumps(data).encode("utf-8")
    last_err: Optional[Exception] = None
    for attempt in range(1, retries + 1):
        with _REQ_LOCK:
            wait = MIN_INTERVAL - (time.time() - _LAST_REQ)
            if wait > 0:
                time.sleep(wait)
            _LAST_REQ = time.time()
        req = urllib.request.Request(
            url,
            data=body,
            method=method,
            headers={
                "User-Agent": UA,
                "Accept": "application/json, text/plain, */*",
                "Accept-Language": "ko-KR,ko;q=0.9,en;q=0.8",
                "Referer": "https://tour.yanolja.com/tna/products?keyword=%EB%B2%A0%ED%8A%B8%EB%82%A8",
                "Origin": "https://tour.yanolja.com",
                "Content-Type": "application/json" if body else "text/plain",
            },
        )
        try:
            with urllib.request.urlopen(req, context=CTX, timeout=30) as resp:
                raw = resp.read()
            if not raw:
                return None
            return json.loads(raw.decode("utf-8"))
        except urllib.error.HTTPError as e:
            last_err = e
            payload = e.read() if hasattr(e, "read") else b""
            if e.code in (400, 404):
                try:
                    return json.loads(payload.decode("utf-8"))
                except Exception:
                    return {"error": {"code": e.code, "message": payload.decode("utf-8", "ignore")}}
            time.sleep(0.8 * attempt + random.random() * 0.4)
        except Exception as e:
            last_err = e
            time.sleep(0.8 * attempt + random.random() * 0.4)
    raise RuntimeError(f"request failed {url} err={last_err}")


def flatten_geotags(geotags: Any) -> List[dict]:
    out: List[dict] = []
    if not geotags:
        return out
    for chain in geotags:
        if isinstance(chain, list):
            out.extend([g for g in chain if isinstance(g, dict)])
        elif isinstance(chain, dict):
            out.append(chain)
    return out


def is_vietnam(product: dict) -> bool:
    tags = flatten_geotags(product.get("geotags"))
    names = {t.get("name") for t in tags}
    ids = {t.get("id") for t in tags}
    if VN_ISO1 in ids or "베트남" in names:
        return True
    # 도시명만 있고 국가 태그가 빠진 경우
    vn_cities = set(SEED_CITIES) | set(CITY_ZH)
    if names & vn_cities:
        return True
    return False


def cities_of(product: dict) -> List[str]:
    tags = flatten_geotags(product.get("geotags"))
    cities = []
    for t in tags:
        if t.get("type") in ("TRIPLE_CITY", "ISO2") and t.get("name"):
            n = t["name"]
            if n not in cities and n != "베트남":
                cities.append(n)
    if not cities:
        for t in tags:
            n = t.get("name")
            if n and n not in ("베트남", "아시아") and n not in cities:
                cities.append(n)
    return cities


def category_names(product: dict) -> List[str]:
    names: List[str] = []

    def walk(node: Any) -> None:
        if isinstance(node, list):
            for x in node:
                walk(x)
            return
        if not isinstance(node, dict):
            return
        n = node.get("name") or node.get("value")
        if n and n not in names:
            names.append(n)
        walk(node.get("child") or node.get("children") or [])

    walk(product.get("categories") or [])
    return names


def text_of(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    if isinstance(value, dict):
        if "source" in value and isinstance(value["source"], str):
            return value["source"].strip()
        vals = value.get("values") or []
        for v in vals:
            if isinstance(v, dict) and v.get("languageCode") == "KO":
                return str(v.get("value") or "").strip()
        if vals and isinstance(vals[0], dict):
            return str(vals[0].get("value") or "").strip()
        for k in ("value", "name", "headline", "highlight", "description"):
            if value.get(k):
                return text_of(value.get(k))
        return ""
    if isinstance(value, list):
        parts = [text_of(x) for x in value]
        return "\n".join([p for p in parts if p])
    return str(value).strip()


TICKET_POS = re.compile(r"입장권|티켓")
TICKET_NEG = re.compile(
    r"마사지|스파|골프|라운지|패스트\s*트랙|패스트트랙|비자|유심|이심|esim|와이파이|"
    r"식사권|쿠킹|네일|페리|투어|조인|가이드|왕복\s*차량|호텔\s*픽업|단독\s*차량|"
    r"단독투어|일일|반일|스노클링|호핑|씨워킹|오토바이|리무진\s*버스|버스 티켓|"
    r"스피드\s*보트|5성급.*크루즈|앰버서더 크루즈",
    re.I,
)
TRANSFER_POS = re.compile(
    r"픽업|샌딩|센딩|공항\s*이동|이동 서비스|이동 차량|단독차량|프라이빗 차량|"
    r"프라이빗차량|셔틀\s*버스|셔틀버스|픽업샌딩|픽업/샌딩|픽업&샌딩|픽업\s*or\s*샌딩|"
    r"렌터카.*(?:픽업|샌딩)|(?:픽업|샌딩).*렌터카",
    re.I,
)
TRANSFER_NEG = re.compile(
    r"투어|스노클링|호핑|쿠킹|씨워킹|오토바이|골든브릿지|가이드|"
    r"일일|반일|조인|마사지|스파|입장권|티켓|시티투어|2층버스|교통패스|"
    r"심카드|esim|이심|글램핑|랜턴 만들기|0\.5박|캐디 골프",
    re.I,
)


def classify(product: dict) -> Optional[str]:
    name = (product.get("name") or "").strip()
    cats = " ".join(category_names(product))
    types = " ".join(
        [
            (t.get("value") or "") + " " + (t.get("name") or "")
            for t in (product.get("types") or [])
            if isinstance(t, dict)
        ]
    )
    is_ticket_cat = "티켓" in cats or "입장권" in cats or "TICKET_PASS" in types
    is_transfer_cat = "교통" in cats or "픽업" in cats or "이동" in cats

    if TRANSFER_POS.search(name) and not TRANSFER_NEG.search(name):
        if is_transfer_cat or re.search(r"공항|호텔|시내|편도|왕복|차량|셔틀", name):
            return "순픽업"

    if TICKET_POS.search(name) and not TICKET_NEG.search(name):
        if is_ticket_cat or "TICKET_PASS" in types or re.search(r"입장권|QR", name):
            return "단입장권"
    return None


def search_page(params: Dict[str, Any]) -> dict:
    qs = urllib.parse.urlencode(params, doseq=True, safe=",")
    url = f"{BASE}/tna-product/externals/products?{qs}"
    data = request_json(url) or {}
    return data


def paginate(label: str, base_params: Dict[str, Any], size: int = 50) -> List[dict]:
    items: List[dict] = []
    page = 1
    total_pages = 1
    while page <= total_pages and page <= 40:
        params = dict(base_params)
        params.update({"page": page, "size": size})
        data = search_page(params)
        if not isinstance(data, dict) or data.get("error"):
            log(f"  ! {label} page {page} error={data}")
            break
        body = data.get("body") or []
        pg = data.get("page") or {}
        total_pages = int(pg.get("totalPages") or 1)
        total_el = pg.get("totalElements")
        log(f"  {label} p{page}/{total_pages} +{len(body)} (total={total_el})")
        items.extend(body)
        if not body:
            break
        page += 1
    return items


def merge_products(dst: Dict[str, dict], rows: Iterable[dict]) -> int:
    added = 0
    for p in rows:
        pid = p.get("id")
        if not pid:
            continue
        if not is_vietnam(p):
            continue
        if pid not in dst:
            dst[pid] = p
            added += 1
        else:
            # 카테고리/가격이 더 풍부한 쪽 유지
            old = dst[pid]
            if not old.get("price") and p.get("price"):
                dst[pid] = p
    return added


def collect_listings() -> Dict[str, dict]:
    DATA.mkdir(parents=True, exist_ok=True)
    products: Dict[str, dict] = {}
    if LISTING_PATH.exists():
        try:
            products = {p["id"]: p for p in json.loads(LISTING_PATH.read_text("utf-8")) if p.get("id")}
            log(f"resume listings {len(products)}")
        except Exception:
            products = {}
    if products and len(products) >= 40:
        log("use cached listings (delete data/listings.json to refresh)")
        return products

    # 1) 키워드 베트남으로 도시 ID 보강
    seed = search_page({"keyword": "베트남", "size": 10, "page": 1})
    city_map = dict(SEED_CITIES)
    for country in ((seed.get("filters") or {}).get("countries") or []):
        if country.get("name") != "베트남":
            continue
        for c in country.get("cities") or []:
            if c.get("value") and c.get("key"):
                city_map[c["value"]] = c["key"]
    log(f"vietnam cities: {len(city_map)} {list(city_map)}")

    # 2) 도시별 × 카테고리 (cityIds 콤마 결합은 OR가 아닐 수 있어 도시 단위로 수집)
    for cname, cid in city_map.items():
        for cat_id, cat_label in ((TICKET_CAT, "ticket"), (TRANSFER_CAT, "transfer")):
            added = merge_products(
                products,
                paginate(
                    f"{cname}/{cat_label}",
                    {"cityIds": cid, "categoryIds": cat_id},
                ),
            )
            log(f"  {cname}/{cat_label} +{added}, unique={len(products)}")

    # 3) 키워드 보완
    extra_kws = [
        "베트남 입장권",
        "베트남 티켓",
        "베트남 픽업",
        "베트남 샌딩",
        "베트남 공항픽업",
        "다낭 입장권",
        "다낭 공항픽업",
        "하노이 공항픽업",
        "호치민 공항픽업",
        "나트랑 공항픽업",
        "푸꾸옥 입장권",
        "바나힐 입장권",
        "빈원더스 입장권",
        "호이안 메모리즈",
    ]
    for kw in extra_kws:
        added = merge_products(products, paginate(f"kw:{kw}", {"keyword": kw}))
        log(f"  kw {kw} +{added}, unique={len(products)}")

    rows = list(products.values())
    LISTING_PATH.write_text(json.dumps(rows, ensure_ascii=False), encoding="utf-8")
    log(f"saved listings {len(rows)} -> {LISTING_PATH}")
    return products


def fetch_detail(pid: str) -> Optional[dict]:
    DETAIL_DIR.mkdir(parents=True, exist_ok=True)
    path = DETAIL_DIR / f"{pid}.json"
    if path.exists():
        try:
            return json.loads(path.read_text("utf-8"))
        except Exception:
            pass
    qs = urllib.parse.urlencode(
        {
            "startDate": START_DATE,
            "endDate": END_DATE,
            "withReview": "true",
            "withItem": "true",
            "withContentAsset": "true",
            "withSupplierNotice": "true",
        }
    )
    url = f"{BASE}/tna-product/externals/products/{pid}?{qs}"
    data = request_json(url)
    body = (data or {}).get("body") if isinstance(data, dict) else None
    if not body:
        return None
    path.write_text(json.dumps(body, ensure_ascii=False), encoding="utf-8")
    return body


def fetch_item_prices(pid: str, item_id: str) -> Optional[dict]:
    ITEM_DIR.mkdir(parents=True, exist_ok=True)
    path = ITEM_DIR / f"{item_id}.json"
    if path.exists():
        try:
            return json.loads(path.read_text("utf-8"))
        except Exception:
            pass
    qs = urllib.parse.urlencode(
        {
            "startDate": START_DATE,
            "endDate": END_DATE,
            "date": START_DATE,
        }
    )
    url = f"{BASE}/tna-product/externals/v2/products/{pid}/items/{item_id}?{qs}"
    data = request_json(url)
    body = (data or {}).get("body") if isinstance(data, dict) else None
    if not body:
        return None
    path.write_text(json.dumps(body, ensure_ascii=False), encoding="utf-8")
    return body


def price_tuple(price: Optional[dict]) -> Tuple[Optional[int], Optional[int], Optional[int], Optional[int]]:
    if not isinstance(price, dict):
        return None, None, None, None
    display = price.get("display")
    sales = price.get("sales")
    coupon = price.get("appliedCoupon")
    rate = price.get("totalDiscountRate")
    if rate is None:
        rate = price.get("discountRate")
    return display, sales, coupon, rate


def option_price_row(opt: dict) -> Tuple[Optional[int], Optional[int], Optional[int], Optional[int]]:
    prices = opt.get("prices") or []
    if prices and isinstance(prices[0], dict):
        p0 = prices[0]
        nested = p0.get("price") if isinstance(p0.get("price"), dict) else p0
        return price_tuple(nested if "display" in nested else p0)
    return price_tuple(opt.get("price") if isinstance(opt.get("price"), dict) else None)


def guide_langs(product: dict) -> str:
    langs = product.get("guideLanguages") or product.get("languages") or []
    names = []
    for x in langs:
        if isinstance(x, dict):
            names.append(x.get("name") or x.get("value") or "")
        else:
            names.append(str(x))
    return ", ".join([n for n in names if n])


def bool_kr(v: Any) -> str:
    if v is True:
        return "Y"
    if v is False:
        return "N"
    return ""


def supplier_name(detail: Optional[dict], listing: dict) -> str:
    if detail:
        sp = detail.get("supplierPartner") or {}
        if isinstance(sp, dict) and sp.get("name"):
            return sp["name"]
        s = detail.get("supplier")
        if isinstance(s, dict):
            return s.get("name") or s.get("supplierPartnerCode") or ""
        if isinstance(s, str):
            return s
    return ""


def review_info(detail: Optional[dict], listing: dict) -> Tuple[Optional[float], Optional[int]]:
    src = detail or listing
    rev = src.get("review") if isinstance(src, dict) else None
    if isinstance(rev, dict):
        return rev.get("totalRating"), rev.get("totalCount")
    return None, None


HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
TICKET_FILL = PatternFill("solid", fgColor="D6EAF8")
TRANSFER_FILL = PatternFill("solid", fgColor="D5F5E3")
WRAP = Alignment(wrap_text=True, vertical="center")
THIN = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)


def style_header(ws, ncols: int) -> None:
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(ncols)}1"
    ws.row_dimensions[1].height = 28
    for col in range(1, ncols + 1):
        cell = ws.cell(1, col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def autosize(ws, max_width: int = 48) -> None:
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        length = 0
        for cell in col[:80]:
            val = "" if cell.value is None else str(cell.value)
            length = max(length, min(len(val), max_width))
        ws.column_dimensions[letter].width = min(max(12, length + 2), max_width)


def write_rows(ws, headers: List[str], rows: List[List[Any]], fill: PatternFill) -> None:
    ws.append(headers)
    style_header(ws, len(headers))
    for r in rows:
        ws.append(r)
        for col in range(1, len(headers) + 1):
            cell = ws.cell(ws.max_row, col)
            cell.alignment = WRAP
            cell.border = THIN
            if col <= 2:
                cell.fill = fill
    autosize(ws)
    if rows:
        tab = Table(
            displayName=re.sub(r"[^A-Za-z0-9]", "", ws.title)[:20] + "Tbl",
            ref=f"A1:{get_column_letter(len(headers))}{ws.max_row}",
        )
        tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
        try:
            ws.add_table(tab)
        except Exception:
            pass


def build_product_row(kind: str, listing: dict, detail: Optional[dict]) -> List[Any]:
    src = dict(listing)
    if detail:
        src.update({k: v for k, v in detail.items() if v not in (None, [], {}, "")})
        # listing 가격이 없을 때만 detail 가격 사용
        if listing.get("price"):
            src["price"] = listing["price"]
    name_ko = src.get("name") or listing.get("name") or ""
    name_zh = translate_ko(name_ko)
    cats = category_names(src) or category_names(listing)
    cat_ko, cat_zh = translate_list(cats)
    cities = cities_of(src) or cities_of(listing)
    city_ko, city_zh = translate_list(cities)
    display, sales, coupon, rate = price_tuple(src.get("price") or listing.get("price"))
    rating, rcount = review_info(detail, listing)
    types = src.get("types") or listing.get("types") or []
    type_ko = ", ".join([(t.get("name") or t.get("value") or "") for t in types if isinstance(t, dict)])
    type_zh = translate_ko(type_ko)
    point = ""
    intro = ""
    if detail:
        point = text_of(detail.get("point"))
        intro = text_of((detail.get("introduction") or {}).get("description") if isinstance(detail.get("introduction"), dict) else detail.get("introduction"))
    thumb = ""
    th = listing.get("thumbnail") or {}
    if isinstance(th, dict):
        url = th.get("url") or {}
        thumb = url.get("large") or url.get("full") or url.get("original") or ""
    badges = []
    for b in listing.get("badges") or []:
        if isinstance(b, dict):
            badges.append(b.get("name") or b.get("value") or "")
        else:
            badges.append(str(b))
    cancel = ",".join(listing.get("cancellationTypes") or src.get("cancellationTypes") or [])
    items = (detail or {}).get("items") or []
    item_names = [text_of(it.get("name")) for it in items if isinstance(it, dict)]
    return [
        kind,
        translate_ko(kind),
        listing.get("id"),
        name_ko,
        name_zh,
        cat_ko,
        cat_zh,
        city_ko,
        city_zh,
        type_ko,
        type_zh,
        display,
        sales,
        coupon,
        rate,
        rating,
        rcount,
        bool_kr(src.get("hasInstantConfirmation") if src.get("hasInstantConfirmation") is not None else listing.get("hasInstantConfirmation")),
        bool_kr(src.get("hasInstantVoucher") if src.get("hasInstantVoucher") is not None else listing.get("hasInstantVoucher")),
        guide_langs(src) or guide_langs(listing),
        cancel,
        supplier_name(detail, listing),
        " / ".join([x for x in item_names if x]),
        " / ".join([translate_ko(x) for x in item_names if x]),
        ", ".join([b for b in badges if b]),
        point,
        translate_ko(point) if point else "",
        (intro or "")[:800],
        translate_ko(intro)[:800] if intro else "",
        PRODUCT_URL.format(id=listing.get("id")),
        thumb,
        START_DATE,
    ]


PRODUCT_HEADERS = [
    "상품구분(한)",
    "상품구분(중)",
    "상품ID",
    "상품명(한)",
    "상품명(중)",
    "카테고리(한)",
    "카테고리(중)",
    "도시(한)",
    "도시(중)",
    "상품타입(한)",
    "상품타입(중)",
    "정상가(KRW)",
    "판매가(KRW)",
    "쿠폰적용가(KRW)",
    "총할인율(%)",
    "평점",
    "리뷰수",
    "즉시확정",
    "즉시바우처",
    "가이드언어",
    "취소기준",
    "판매자",
    "세트/패키지(한)",
    "세트/패키지(중)",
    "뱃지",
    "매력포인트(한)",
    "매력포인트(중)",
    "소개(한)",
    "소개(중)",
    "상품URL",
    "썸네일",
    "가격기준일",
]

OPTION_HEADERS = [
    "상품구분(한)",
    "상품구분(중)",
    "상품ID",
    "상품명(한)",
    "상품명(중)",
    "패키지/아이템ID",
    "패키지명(한)",
    "패키지명(중)",
    "옵션ID",
    "옵션명(한)",
    "옵션명(중)",
    "옵션설명(한)",
    "옵션설명(중)",
    "정상가(KRW)",
    "판매가(KRW)",
    "쿠폰적용가(KRW)",
    "할인율(%)",
    "최소인원",
    "최대인원",
    "포함사항(한)",
    "포함사항(중)",
    "불포함(한)",
    "불포함(중)",
    "사용방법(한)",
    "사용방법(중)",
    "가격기준일",
]


def build_option_rows(kind: str, listing: dict, detail: Optional[dict], item_bodies: Dict[str, dict]) -> List[List[Any]]:
    rows: List[List[Any]] = []
    name_ko = listing.get("name") or ""
    name_zh = translate_ko(name_ko)
    items = (detail or {}).get("items") or []
    if not items:
        return rows
    for it in items:
        if not isinstance(it, dict):
            continue
        iid = it.get("id")
        rich = item_bodies.get(iid) or it
        pkg_ko = text_of(rich.get("name") or it.get("name"))
        pkg_zh = translate_ko(pkg_ko)
        inclusions = []
        for inc in rich.get("inclusions") or it.get("inclusions") or []:
            if isinstance(inc, dict):
                inclusions.append(text_of(inc.get("description")) or text_of(inc.get("code")))
            else:
                inclusions.append(str(inc))
        inc_ko = "\n".join([x for x in inclusions if x])
        exclusion = text_of(rich.get("exclusion") or it.get("exclusion"))
        usage = ""
        add = rich.get("additional") or it.get("additional") or {}
        if isinstance(add, dict):
            usage = text_of(add.get("usage"))
        opts = rich.get("options") or it.get("options") or []
        if not opts:
            d, s, c, r = price_tuple(rich.get("price"))
            rows.append(
                [
                    kind,
                    translate_ko(kind),
                    listing.get("id"),
                    name_ko,
                    name_zh,
                    iid,
                    pkg_ko,
                    pkg_zh,
                    "",
                    "(기본)",
                    "(默认)",
                    "",
                    "",
                    d,
                    s,
                    c,
                    r,
                    (rich.get("booking") or {}).get("minPax"),
                    (rich.get("booking") or {}).get("maxPax"),
                    inc_ko,
                    translate_ko(inc_ko),
                    exclusion,
                    translate_ko(exclusion),
                    usage,
                    translate_ko(usage),
                    START_DATE,
                ]
            )
            continue
        for opt in opts:
            if not isinstance(opt, dict):
                continue
            d, s, c, r = option_price_row(opt)
            opt_ko = text_of(opt.get("name"))
            desc_ko = text_of(opt.get("description"))
            book = opt.get("booking") or {}
            rows.append(
                [
                    kind,
                    translate_ko(kind),
                    listing.get("id"),
                    name_ko,
                    name_zh,
                    iid,
                    pkg_ko,
                    pkg_zh,
                    opt.get("id"),
                    opt_ko,
                    translate_ko(opt_ko),
                    desc_ko,
                    translate_ko(desc_ko),
                    d,
                    s,
                    c,
                    r,
                    book.get("minPax"),
                    book.get("maxPax"),
                    inc_ko,
                    translate_ko(inc_ko),
                    exclusion,
                    translate_ko(exclusion),
                    usage,
                    translate_ko(usage),
                    START_DATE,
                ]
            )
    return rows


def write_xlsx(classified: Dict[str, List[Tuple[dict, Optional[dict], Dict[str, dict]]]]) -> None:
    from export_xlsx import write_xlsx as _write

    _write(classified)
    log(f"wrote {XLSX_PATH}")



def enrich_classified(products: Dict[str, dict]) -> Dict[str, List[Tuple[dict, Optional[dict], Dict[str, dict]]]]:
    buckets: Dict[str, List[dict]] = {"단입장권": [], "순픽업": []}
    skipped = 0
    for p in products.values():
        kind = classify(p)
        if not kind:
            skipped += 1
            continue
        buckets[kind].append(p)
    log(f"classified tickets={len(buckets['단입장권'])} transfers={len(buckets['순픽업'])} skipped={skipped}")

    classified: Dict[str, List[Tuple[dict, Optional[dict], Dict[str, dict]]]] = {
        "단입장권": [],
        "순픽업": [],
    }

    all_targets = [(k, p) for k, rows in buckets.items() for p in rows]
    details: Dict[str, Optional[dict]] = {}

    def _detail(pid: str) -> Tuple[str, Optional[dict]]:
        try:
            return pid, fetch_detail(pid)
        except Exception as e:
            log(f"  detail fail {pid} {e}")
            return pid, None

    log(f"fetch details n={len(all_targets)}")
    with ThreadPoolExecutor(max_workers=4) as ex:
        futs = [ex.submit(_detail, p["id"]) for _, p in all_targets]
        done = 0
        for fut in as_completed(futs):
            pid, body = fut.result()
            details[pid] = body
            done += 1
            if done % 20 == 0 or done == len(futs):
                log(f"  details {done}/{len(futs)}")

    # 분류를 detail 기준으로 한 번 더 정제
    item_jobs: List[Tuple[str, str, str]] = []  # kind, pid, item_id
    for kind, p in all_targets:
        d = details.get(p["id"])
        merged = dict(p)
        if d:
            # types/categories from detail are more accurate
            if d.get("types"):
                merged["types"] = d["types"]
            if d.get("categories"):
                merged["categories"] = d["categories"]
            if d.get("name"):
                merged["name"] = d["name"] if isinstance(d["name"], str) else text_of(d["name"])
        kind2 = classify(merged) or kind
        items = {}
        classified.setdefault(kind2, [])
        classified[kind2].append((p, d, items))
        if d:
            for it in d.get("items") or []:
                if isinstance(it, dict) and it.get("id"):
                    item_jobs.append((kind2, p["id"], it["id"]))

    log(f"fetch item prices n={len(item_jobs)}")
    item_map: Dict[str, dict] = {}

    def _item(pid: str, iid: str) -> Tuple[str, Optional[dict]]:
        try:
            return iid, fetch_item_prices(pid, iid)
        except Exception as e:
            log(f"  item fail {iid} {e}")
            return iid, None

    with ThreadPoolExecutor(max_workers=4) as ex:
        futs = [ex.submit(_item, pid, iid) for _, pid, iid in item_jobs]
        done = 0
        for fut in as_completed(futs):
            iid, body = fut.result()
            if body:
                item_map[iid] = body
            done += 1
            if done % 30 == 0 or done == len(futs):
                log(f"  items {done}/{len(futs)}")

    # attach items
    out: Dict[str, List[Tuple[dict, Optional[dict], Dict[str, dict]]]] = {"단입장권": [], "순픽업": []}
    for kind, rows in classified.items():
        if kind not in out:
            continue
        for listing, detail, _ in rows:
            attached = {}
            if detail:
                for it in detail.get("items") or []:
                    iid = it.get("id")
                    if iid and iid in item_map:
                        attached[iid] = item_map[iid]
            out[kind].append((listing, detail, attached))
    # 이름순
    for kind in out:
        out[kind].sort(key=lambda x: ((cities_of(x[0]) or [""])[0], x[0].get("name") or ""))
    return out


def load_cached_classified() -> Dict[str, List[Tuple[dict, Optional[dict], Dict[str, dict]]]]:
    products = {p["id"]: p for p in json.loads(LISTING_PATH.read_text("utf-8")) if p.get("id")}
    classified: Dict[str, List[Tuple[dict, Optional[dict], Dict[str, dict]]]] = {
        "단입장권": [],
        "순픽업": [],
    }
    for p in products.values():
        d = None
        dp = DETAIL_DIR / f"{p['id']}.json"
        if dp.exists():
            try:
                d = json.loads(dp.read_text("utf-8"))
            except Exception:
                d = None
        merged = dict(p)
        if d:
            if isinstance(d.get("name"), str):
                merged["name"] = d["name"]
            elif d.get("name"):
                merged["name"] = text_of(d.get("name"))
            if d.get("types"):
                merged["types"] = d["types"]
            if d.get("categories"):
                merged["categories"] = d["categories"]
        kind = classify(merged)
        if not kind:
            continue
        attached: Dict[str, dict] = {}
        if d:
            for it in d.get("items") or []:
                iid = it.get("id") if isinstance(it, dict) else None
                ip = ITEM_DIR / f"{iid}.json" if iid else None
                if ip and ip.exists():
                    try:
                        attached[iid] = json.loads(ip.read_text("utf-8"))
                    except Exception:
                        pass
        classified[kind].append((p, d, attached))
    for kind in classified:
        classified[kind].sort(key=lambda x: ((cities_of(x[0]) or [""])[0], x[0].get("name") or ""))
        log(f"cached {kind}={len(classified[kind])}")
    return classified


def main() -> None:
    import sys

    DATA.mkdir(parents=True, exist_ok=True)
    log(f"START_DATE={START_DATE} END_DATE={END_DATE}")
    if "--xlsx-only" in sys.argv:
        classified = load_cached_classified()
        write_xlsx(classified)
        log("done xlsx-only")
        return
    products = collect_listings()
    classified = enrich_classified(products)
    write_xlsx(classified)
    log("done")


if __name__ == "__main__":
    main()
