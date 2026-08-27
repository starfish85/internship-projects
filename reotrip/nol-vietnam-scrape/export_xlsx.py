# -*- coding: utf-8 -*-
"""按 vietnam_tickets_transfers.xlsx 的字段与样式导出 NOL 韩中对照表。"""

from __future__ import annotations

import json
import re
import time
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Side, Border
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from ko_zh import translate_ko, translate_list
from scrape_nol_vietnam import (
    PRODUCT_URL,
    XLSX_PATH,
    category_names,
    cities_of,
    option_price_row,
    price_tuple,
    text_of,
)

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(name="宋体", bold=True, color="FFFFFF", size=11)
DATA_FONT = Font(name="宋体", size=11)
HEADER_AL = Alignment(wrap_text=True, vertical="center")
DATA_AL = Alignment(wrap_text=True, vertical="top")
THIN = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)

PRODUCT_HEADERS = [
    "分类",
    "产品ID",
    "链接",
    "城市（韩）",
    "城市（中）",
    "标题（韩）",
    "标题（中）",
    "副标题（韩）",
    "副标题（中）",
    "格式（韩）",
    "格式（中）",
    "成团类型（韩）",
    "成团类型（中）",
    "出行方式（韩）",
    "出行方式（中）",
    "时长(小时)",
    "最多人数",
    "可带儿童",
    "语言（韩）",
    "语言（中）",
    "价格",
    "货币",
    "计价单位（韩）",
    "计价单位（中）",
    "价格展示（韩）",
    "已下单人数",
    "前端口径",
    "评论数",
    "评分",
    "评分20",
    "热度",
    "即时确认",
    "状态",
    "集合点（韩）",
    "集合点（中）",
    "简介（韩）",
    "简介（中）",
    "描述（韩）",
    "描述（中）",
    "套餐数",
    "封面图",
    "组织者",
    "组织者订单数",
    "组织者接待人数",
    "组织者已付款人数",
    "组织者评分",
    "标签（韩）",
    "可订日期数",
    "抓取时间",
]

OPTION_HEADERS = [
    "分类",
    "产品ID",
    "产品标题（韩）",
    "产品标题（中）",
    "套餐ID",
    "套餐名（韩）",
    "套餐名（中）",
    "套餐说明（韩）",
    "套餐说明（中）",
    "计价类型（韩）",
    "计价类型（中）",
    "价格",
    "货币",
    "展示（韩）",
    "单位（韩）",
    "最少数量",
    "最多数量",
    "提前截止(小时)",
]

TRAFFIC_HEADERS = [
    "分类",
    "产品ID",
    "标题（韩）",
    "标题（中）",
    "城市（中）",
    "已下单人数",
    "前端文案（韩）",
    "前端文案（中）",
    "评论数",
    "评分",
    "评分20",
    "热度",
    "组织者",
    "组织者累计订单",
    "组织者累计接待",
    "组织者已付款人数",
    "组织者评分",
    "组织者评论数",
    "链接",
]

FIELD_ROWS = [
    ["分类", "本地规则", "单门票：景点入场券，不含跟车/导游团。纯接送：机场接送、包车或班车点对点，不含「含接送的一日游」。"],
    ["标题 / 副标题 / 描述", "NOL 产品详情接口", "韩语原文 + 中文对照，便于阅读。"],
    ["套餐", "NOL 套餐/票种接口", "前端预订套餐、票种、车型、区域。无拆分选项时按套餐价记一行。"],
    ["价格", "列表价 / 套餐选项价", "韩元（KRW），展示价优先取销售价。抓取时间见产品对照最后一列。"],
    ["评论数 / 评分", "产品详情评论", "NOL、TRIPLE、INTERPARK 综合评论数与评分（5 分制）。评分20 取 NOL 平台评分，没有则用综合分。"],
    ["已下单人数 / 热度", "NOL 前端未公开", "韩国 NOL 产品页不展示已下单人数和热度，本列留空。"],
    ["组织者", "销售商信息", "销售商名称。NOL 无组织者累计订单、接待人数，对应列留空。"],
    ["即时确认", "产品详情", "是 / 否。"],
    ["集合点", "产品地点", "接送集合点或景点地址。"],
    ["可订日期数", "可订日期列表", "接口返回的可订日期数量。"],
]

PRODUCT_WIDTHS = {
    "A": 12, "B": 22, "C": 42, "D": 22, "E": 13, "F": 40, "G": 40, "H": 36, "I": 36,
    "J": 18, "K": 16, "L": 16, "M": 20, "N": 14, "O": 14, "P": 12, "Q": 12, "R": 12,
    "S": 16, "T": 16, "U": 12, "V": 10, "W": 16, "X": 14, "Y": 22, "Z": 18,
    "AA": 22, "AB": 10, "AC": 10, "AD": 10, "AE": 14, "AF": 12, "AG": 12,
    "AH": 28, "AI": 28, "AJ": 50, "AK": 50, "AL": 55, "AM": 55, "AN": 10, "AO": 40,
    "AP": 22, "AQ": 14, "AR": 16, "AS": 16, "AT": 12, "AU": 36, "AV": 12, "AW": 20,
}
OPTION_WIDTHS = {
    "A": 12, "B": 22, "C": 42, "D": 42, "E": 22, "F": 32, "G": 32, "H": 40, "I": 40,
    "J": 18, "K": 14, "L": 12, "M": 10, "N": 18, "O": 12, "P": 12, "Q": 12, "R": 14,
}
TRAFFIC_WIDTHS = {
    "A": 12, "B": 22, "C": 42, "D": 42, "E": 13, "F": 18, "G": 16, "H": 22, "I": 10,
    "J": 10, "K": 10, "L": 14, "M": 20, "N": 16, "O": 16, "P": 18, "Q": 12, "R": 12, "S": 42,
}

LANG_KO = {"KO": "한국어", "EN": "영어", "VI": "베트남어", "JA": "일본어", "ZH": "중국어", "CN": "중국어"}
LANG_ZH = {
    "KO": "韩语",
    "EN": "英语",
    "VI": "越南语",
    "JA": "日语",
    "ZH": "中文",
    "CN": "中文",
    "한국어": "韩语",
    "영어": "英语",
    "베트남어": "越南语",
    "일본어": "日语",
    "중국어": "中文",
    "중국어(간체)": "中文(简体)",
    "중국어(번체)": "中文(繁体)",
    "태국어": "泰语",
    "인도네시아어": "印尼语",
    "스페인어": "西班牙语",
    "프랑스어": "法语",
    "독일어": "德语",
    "러시아어": "俄语",
}

KIND_ZH = {"단입장권": "单门票", "순픽업": "纯接送"}


def _yn(v: Any) -> str:
    if v is True:
        return "是"
    if v is False:
        return "否"
    return ""


def _clip(text: str, n: int = 2000) -> str:
    text = (text or "").strip()
    if len(text) <= n:
        return text
    return text[: n - 1] + "…"


def _src(listing: dict, detail: Optional[dict]) -> dict:
    out = dict(listing)
    if detail:
        for k, v in detail.items():
            if v not in (None, [], {}, ""):
                out[k] = v
        if listing.get("price"):
            out["price"] = listing["price"]
        if isinstance(detail.get("name"), str):
            out["name"] = detail["name"]
        elif detail.get("name"):
            out["name"] = text_of(detail.get("name"))
    return out


def _cover(listing: dict, detail: Optional[dict]) -> str:
    th = listing.get("thumbnail") or {}
    if isinstance(th, dict):
        url = th.get("url") or {}
        if isinstance(url, dict):
            return url.get("large") or url.get("full") or url.get("original") or ""
    if detail:
        for img in detail.get("images") or []:
            if not isinstance(img, dict):
                continue
            url = img.get("url") or {}
            if isinstance(url, dict) and (url.get("large") or url.get("full")):
                return url.get("large") or url.get("full")
    return ""


def _intro(detail: Optional[dict]) -> Tuple[str, str]:
    if not detail:
        return "", ""
    point = detail.get("point") or {}
    headline = text_of(point.get("headline") if isinstance(point, dict) else "")
    highlight = text_of(point.get("highlight") if isinstance(point, dict) else "")
    tag = "\n".join([x for x in [headline, highlight] if x]).strip()
    intro = detail.get("introduction") or {}
    desc = text_of(intro.get("description") if isinstance(intro, dict) else intro)
    if not desc:
        desc = text_of(detail.get("description"))
    return tag, desc


def _meeting(detail: Optional[dict]) -> str:
    if not detail:
        return ""
    places = detail.get("places") or []
    meeting = []
    product = []
    for p in places:
        if not isinstance(p, dict):
            continue
        ptype = ((p.get("type") or {}).get("value") if isinstance(p.get("type"), dict) else "") or ""
        name = text_of(p.get("name"))
        addr = ""
        loc = p.get("location") or {}
        if isinstance(loc, dict):
            addr = loc.get("address") or ""
        label = " / ".join([x for x in [name, addr] if x])
        if not label:
            continue
        if ptype == "MEETING":
            meeting.append(label)
        else:
            product.append(label)
    return " | ".join(meeting or product)


def _languages(src: dict) -> Tuple[str, str]:
    names_ko = []
    names_zh = []
    langs = src.get("languages") or []
    if langs and isinstance(langs[0], dict) and langs[0].get("code"):
        for x in langs:
            code = (x.get("code") or {}) if isinstance(x.get("code"), dict) else {}
            ko = code.get("name") or code.get("value") or ""
            val = (code.get("value") or "").upper()
            names_ko.append(ko or LANG_KO.get(val, val))
            names_zh.append(LANG_ZH.get(val) or LANG_ZH.get(ko) or translate_ko(ko) or val)
    else:
        for code in src.get("guideLanguages") or []:
            if isinstance(code, dict):
                val = (code.get("value") or code.get("code") or "").upper()
                ko = code.get("name") or LANG_KO.get(val, val)
            else:
                val = str(code).upper()
                ko = LANG_KO.get(val, str(code))
            names_ko.append(ko)
            names_zh.append(LANG_ZH.get(val, translate_ko(ko) or str(code)))
    # unique keep order
    def uniq(seq):
        seen = set()
        out = []
        for x in seq:
            if x and x not in seen:
                seen.add(x)
                out.append(x)
        return out

    ko = ", ".join(uniq(names_ko))
    zh = ", ".join(uniq(names_zh))
    return ko, zh


def _child_ok(detail: Optional[dict], items_map: Dict[str, dict]) -> str:
    names = []
    for it in (detail or {}).get("items") or []:
        names.append(text_of(it.get("name")))
        for opt in it.get("options") or []:
            names.append(text_of(opt.get("name")))
            names.append(text_of(opt.get("description")))
        rich = items_map.get(it.get("id") or "")
        if rich:
            names.append(text_of(rich.get("name")))
            for opt in rich.get("options") or []:
                names.append(text_of(opt.get("name")))
    blob = " ".join(names)
    return "是" if re.search(r"아동|소아|유아|어린이|키\s*\d", blob) else "否"


def _group_and_travel(kind: str, src: dict) -> Tuple[str, str, str, str]:
    name = src.get("name") or ""
    if kind == "단입장권":
        return "private", "自行入园（单门票）", "foot", "自行前往"
    if re.search(r"공용|셔틀|버스", name):
        return "group", "拼车/班车", "car", "车辆"
    return "private", "包车", "car", "车辆"


def _pricing_unit(kind: str, src: dict) -> Tuple[str, str, str]:
    name = src.get("name") or ""
    if kind == "순픽업" or re.search(r"단독차량|차량 1대|대당|프라이빗 차량", name):
        return "대당", "每车", "/대"
    return "인당", "每人", "/인"


def _max_pax(detail: Optional[dict]) -> Optional[int]:
    vals = []
    for it in (detail or {}).get("items") or []:
        book = it.get("booking") or {}
        mx = book.get("maxPax")
        if isinstance(mx, int) and 0 < mx < 9999:
            vals.append(mx)
    return max(vals) if vals else None


def _lead_hours(detail: Optional[dict]) -> Optional[float]:
    hours = []
    for it in (detail or {}).get("items") or []:
        lead = (it.get("booking") or {}).get("averageLeadTime") or {}
        h = (lead.get("days") or 0) * 24 + (lead.get("hours") or 0) + (lead.get("minutes") or 0) / 60
        if h:
            hours.append(h)
    return hours[0] if hours else None


def _usable_count(src: dict) -> int:
    dates = src.get("usableDates") or []
    return len(dates) if isinstance(dates, list) else 0


def _review(src: dict) -> Tuple[Optional[int], Optional[float], Optional[float]]:
    rev = src.get("review") or {}
    if not isinstance(rev, dict):
        return None, None, None
    count = rev.get("totalCount")
    rating = rev.get("totalRating")
    nol = None
    for p in rev.get("platforms") or []:
        if isinstance(p, dict) and p.get("platform") == "NOL" and p.get("count"):
            nol = p.get("rating")
            break
    return count, rating, nol if nol is not None else rating


def _tags(src: dict) -> str:
    parts = category_names(src)
    for b in src.get("badges") or []:
        if isinstance(b, dict):
            parts.append(b.get("name") or b.get("value") or "")
        else:
            parts.append(str(b))
    return ", ".join([x for x in parts if x])


def _is_per_group(kind: str, item_name: str, opt_name: str) -> bool:
    blob = f"{item_name} {opt_name}"
    if kind == "순픽업":
        return True
    return bool(re.search(r"인승|차량|대당|SUV|승합|승용|솔라티", blob))


def style_sheet(ws: Worksheet, headers: List[str], widths: Dict[str, float], n_rows: int) -> None:
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 28
    last = get_column_letter(len(headers))
    if n_rows >= 1:
        ws.auto_filter.ref = f"A1:{last}{n_rows}"
    for col, h in enumerate(headers, 1):
        cell = ws.cell(1, col, h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_AL
        letter = get_column_letter(col)
        ws.column_dimensions[letter].width = widths.get(letter, 14)
    for r in range(2, n_rows + 1):
        ws.row_dimensions[r].height = 18
        for c in range(1, len(headers) + 1):
            cell = ws.cell(r, c)
            cell.font = DATA_FONT
            cell.alignment = DATA_AL


def product_row(
    kind: str,
    listing: dict,
    detail: Optional[dict],
    items_map: Dict[str, dict],
    scraped_at: str,
) -> Tuple[List[Any], int]:
    src = _src(listing, detail)
    name_ko = src.get("name") or listing.get("name") or ""
    name_zh = translate_ko(name_ko)
    cities = cities_of(src) or cities_of(listing)
    city_ko, city_zh = translate_list(cities)
    if not city_zh and cities:
        city_zh = translate_ko(cities[0])
    types = src.get("types") or []
    type_ko = ", ".join([(t.get("name") or t.get("value") or "") for t in types if isinstance(t, dict)])
    cat_zh = KIND_ZH.get(kind, kind)
    g_ko, g_zh, tr_ko, tr_zh = _group_and_travel(kind, src)
    unit_ko, unit_zh, unit_suffix = _pricing_unit(kind, src)
    display, sales, coupon, _rate = price_tuple(src.get("price") or listing.get("price"))
    price = sales if sales is not None else display
    lang_ko, lang_zh = _languages(src)
    tag, desc = _intro(detail)
    meet = _meeting(detail)
    n_opts = 0
    items = (detail or {}).get("items") or []
    for it in items:
        rich = items_map.get(it.get("id") or "") or it
        opts = rich.get("options") or it.get("options") or []
        n_opts += max(1, len(opts))
    rcount, rating, rating20 = _review(src)
    cover = _cover(listing, detail)
    supplier = ""
    if isinstance(src.get("supplierPartner"), dict):
        supplier = src["supplierPartner"].get("name") or ""
    price_show = f"{price}원{unit_suffix}" if price is not None else ""
    fmt_ko = type_ko or ("티켓・패스" if kind == "단입장권" else "픽업∙샌딩")
    fmt_zh = "单门票" if kind == "단입장권" else "纯接送"
    row = [
        cat_zh,
        listing.get("id"),
        PRODUCT_URL.format(id=listing.get("id")),
        city_ko,
        city_zh,
        name_ko,
        name_zh,
        _clip(tag, 300),
        _clip(translate_ko(tag), 300) if tag else "",
        fmt_ko,
        fmt_zh,
        g_ko,
        g_zh,
        tr_ko,
        tr_zh,
        None,
        _max_pax(detail),
        _child_ok(detail, items_map),
        lang_ko,
        lang_zh,
        price,
        "KRW",
        unit_ko,
        unit_zh,
        price_show,
        None,
        "NOL未公开 / 前端未展示下单人数",
        rcount,
        rating,
        rating20,
        None,
        _yn(src.get("hasInstantConfirmation")),
        "active" if src.get("saleable") is not False else "inactive",
        meet,
        translate_ko(meet) if meet else "",
        _clip(tag, 1500),
        _clip(translate_ko(tag), 1500) if tag else "",
        _clip(desc, 2500),
        _clip(translate_ko(desc), 2500) if desc else "",
        n_opts,
        cover,
        supplier,
        None,
        None,
        None,
        None,
        _tags(src),
        _usable_count(src),
        scraped_at,
    ]
    return row, n_opts


def option_rows(
    kind: str,
    listing: dict,
    detail: Optional[dict],
    items_map: Dict[str, dict],
) -> List[List[Any]]:
    src = _src(listing, detail)
    name_ko = src.get("name") or listing.get("name") or ""
    name_zh = translate_ko(name_ko)
    cat_zh = KIND_ZH.get(kind, kind)
    rows: List[List[Any]] = []
    items = (detail or {}).get("items") or []
    lead = _lead_hours(detail)
    if not items:
        display, sales, _, _ = price_tuple(src.get("price") or listing.get("price"))
        price = sales if sales is not None else display
        per_group = kind == "순픽업"
        unit = "/대" if per_group else "/인"
        show = f"{price}원{unit}" if price is not None else ""
        rows.append(
            [
                cat_zh,
                listing.get("id"),
                name_ko,
                name_zh,
                listing.get("id"),
                name_ko,
                name_zh,
                "",
                "",
                "per_group_price" if per_group else "per_person_price",
                "按车计价" if per_group else "按人计价",
                price,
                "KRW",
                show,
                unit,
                1,
                None,
                lead,
            ]
        )
        return rows

    for it in items:
        if not isinstance(it, dict):
            continue
        iid = it.get("id")
        rich = items_map.get(iid) or it
        pkg_ko = text_of(rich.get("name") or it.get("name"))
        pkg_zh = translate_ko(pkg_ko)
        desc_ko = text_of(rich.get("description") or it.get("description"))
        opts = rich.get("options") or it.get("options") or []
        book = rich.get("booking") or it.get("booking") or {}
        item_lead = None
        lt = book.get("averageLeadTime") or {}
        if lt:
            item_lead = (lt.get("days") or 0) * 24 + (lt.get("hours") or 0) + (lt.get("minutes") or 0) / 60
        if not opts:
            display, sales, _, _ = price_tuple(rich.get("price") if isinstance(rich.get("price"), dict) else src.get("price"))
            price = sales if sales is not None else display
            per_group = _is_per_group(kind, pkg_ko, "")
            unit = "/대" if per_group else "/인"
            rows.append(
                [
                    cat_zh,
                    listing.get("id"),
                    name_ko,
                    name_zh,
                    iid,
                    pkg_ko,
                    pkg_zh,
                    _clip(desc_ko, 800),
                    _clip(translate_ko(desc_ko), 800) if desc_ko else "",
                    "per_group_price" if per_group else "per_person_price",
                    "按车计价" if per_group else "按人计价",
                    price,
                    "KRW",
                    f"{price}원{unit}" if price is not None else "",
                    unit,
                    book.get("minPax") or 1,
                    book.get("maxPax") if book.get("maxPax") not in (None, 9999) else None,
                    item_lead if item_lead else lead,
                ]
            )
            continue
        for opt in opts:
            if not isinstance(opt, dict):
                continue
            opt_ko = text_of(opt.get("name"))
            opt_desc = text_of(opt.get("description"))
            title_ko = f"{pkg_ko} / {opt_ko}" if opt_ko and opt_ko != pkg_ko else (opt_ko or pkg_ko)
            title_zh = translate_ko(title_ko)
            d, s, _, _ = option_price_row(opt)
            price = s if s is not None else d
            per_group = _is_per_group(kind, pkg_ko, opt_ko)
            unit = "/대" if per_group else "/인"
            explain = opt_desc or desc_ko
            obook = opt.get("booking") or {}
            rows.append(
                [
                    cat_zh,
                    listing.get("id"),
                    name_ko,
                    name_zh,
                    opt.get("id") or iid,
                    title_ko,
                    title_zh,
                    _clip(explain, 800),
                    _clip(translate_ko(explain), 800) if explain else "",
                    "per_group_price" if per_group else "per_person_price",
                    "按车计价" if per_group else "按人计价",
                    price,
                    "KRW",
                    f"{price}원{unit}" if price is not None else "",
                    unit,
                    obook.get("minPax") if obook.get("minPax") not in (None, 0) else 1,
                    obook.get("maxPax") if obook.get("maxPax") not in (None, 9999) else None,
                    item_lead if item_lead else lead,
                ]
            )
    return rows


def write_xlsx(classified: Dict[str, List[Tuple[dict, Optional[dict], Dict[str, dict]]]]) -> None:
    scraped_at = time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime())
    wb = Workbook()

    prod_rows: List[List[Any]] = []
    opt_rows: List[List[Any]] = []
    traffic_rows: List[List[Any]] = []
    for kind in ("단입장권", "순픽업"):
        for listing, detail, items_map in classified.get(kind) or []:
            prow, _n = product_row(kind, listing, detail, items_map, scraped_at)
            prod_rows.append(prow)
            opt_rows.extend(option_rows(kind, listing, detail, items_map))
            traffic_rows.append(
                [
                    prow[0],
                    prow[1],
                    prow[5],
                    prow[6],
                    prow[4],
                    prow[25],
                    "리뷰",
                    "评论数（前端展示）",
                    prow[27],
                    prow[28],
                    prow[29],
                    prow[30],
                    prow[41],
                    None,
                    None,
                    None,
                    None,
                    prow[27],
                    prow[2],
                ]
            )

    ws = wb.active
    ws.title = "产品对照"
    ws.append(PRODUCT_HEADERS)
    for row in prod_rows:
        ws.append(row)
    style_sheet(ws, PRODUCT_HEADERS, PRODUCT_WIDTHS, ws.max_row)

    ws = wb.create_sheet("套餐对照")
    ws.append(OPTION_HEADERS)
    for row in opt_rows:
        ws.append(row)
    style_sheet(ws, OPTION_HEADERS, OPTION_WIDTHS, ws.max_row)

    ws = wb.create_sheet("流量与订单")
    ws.append(TRAFFIC_HEADERS)
    for row in traffic_rows:
        ws.append(row)
    style_sheet(ws, TRAFFIC_HEADERS, TRAFFIC_WIDTHS, ws.max_row)

    ws = wb.create_sheet("字段说明")
    ws.append(["字段", "来源", "说明"])
    for row in FIELD_ROWS:
        ws.append(row)
    style_sheet(ws, ["字段", "来源", "说明"], {"A": 28, "B": 32, "C": 80}, ws.max_row)
    ws.row_dimensions[1].height = 28

    XLSX_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(XLSX_PATH)
