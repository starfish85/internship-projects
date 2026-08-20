#!/usr/bin/env python3
"""Classify Tripster public listing/detail data into xlsx."""
from __future__ import annotations

import json
import re
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

SRC_PRODUCTS = Path("/tmp/tripster-scrape/out/products.json")
SRC_LISTINGS = Path("/tmp/tripster-scrape/out/listings.json")
SRC_FAILS = Path("/tmp/tripster-scrape/out/failures.json")
OUT_XLSX = Path("/Users/mac/景区整理/tripster_transfer_ticket_20260817.xlsx")

# USD base from open.er-api.com 2026-08-17
USD_CNY = 6.748651
TO_CNY = {
    "USD": 6.748651,
    "EUR": 7.809402,
    "THB": 0.203663,
    "TRY": 0.140936,
    "RUB": 0.080885,
    "AED": 1.837618,
    "IDR": 0.00037842,
    "GBP": 8.60,  # fallback unused unless seen
}
FX_DATE = "2026-08-17"
FX_NOTE = "open.er-api.com mid-market, 2026-08-17; 仅估算"

CITY_COUNTRY = {
    "Dubai": "ОАЭ / 阿联酋",
    "Istanbul": "Турция / 土耳其",
    "Bangkok": "Таиланд / 泰国",
    "Phuket": "Таиланд / 泰国",
    "Bali": "Индонезия / 印尼",
}
CITY_CN = {
    "Dubai": "迪拜",
    "Istanbul": "伊斯坦布尔",
    "Bangkok": "曼谷",
    "Phuket": "普吉",
    "Bali": "巴厘岛",
}

TRANSFER_RE = re.compile(
    r"трансфер|трансферы|transfer|transfers|аэропорт.?отел|airport transfer",
    re.I,
)
TICKET_RE = re.compile(
    r"\bбилет|\bбилеты|\bticket|\btickets|skip[-\s]?the[-\s]?line|входн",
    re.I,
)
GUIDE_RE = re.compile(
    r"рассказ(?:ы|ами|ов)? об|"
    r"поделится полезной информацией|"
    r"инфосопровожден|"
    r"аудиоэкскурси|"
    r"аудиогид|"
    r"с гидом|"
    r"гид сопровожд|"
    r"экскурсовод|"
    r"пешеходн(?:ая|ый) экскурси|"
    r"обзорная экскурси|"
    r"авторск(?:ая|ий) экскурси|"
    r"трансфер \+ экскурси|"
    r"трансфер.+обзорн|"
    r"с экскурсией",
    re.I,
)
SIGHTSEEING_RE = re.compile(
    r"за один день увидеть|такая программа|отправимся в два места|"
    r"погуляете и проедете|джип-сафари|вертол[её]т|"
    r"романтическ(?:ий|ие) ужин|ужин посреди|"
    r"hop-?on hop-?off|hop on hop off|"
    r"круиз|на кабриолете вы пронес|"
    r"пляжный (?:день|отдых)|тематический парк|"
    r"знакомство со? |обзорн",
    re.I,
)
TOUR_EXCLUDE_RE = re.compile(
    r"обзорн(?:ая|ые) экскурси|авторск(?:ая|ие) экскурси|"
    r"гастрономическ|фотосесс|мастер-класс|дайвинг|"
    r"snorkeling|снорклинг|квадроцикл|багги|"
    r"сафари по|джип-сафари|вертол[её]тн",
    re.I,
)
ADMISSION_TITLE_RE = re.compile(
    r"аквариум|аквапарк|шоу |кабаре|смотров|музей|"
    r"парк (?:ferrari|lego|warner|motion|img)|"
    r"zipline|ziplajn|xline|сити пасс|city pass|"
    r"обсерватор|вход",
    re.I,
)

COLS = [
    "产品ID",
    "产品URL",
    "产品标题（原文）",
    "城市",
    "国家",
    "分类判定",
    "判定依据",
    "产品类型/标签",
    "时长",
    "起始价",
    "币种",
    "起始价_人民币",
    "汇率",
    "汇率日期",
    "计价单位",
    "是否含接送",
    "是否含门票",
    "是否含导游",
    "集合方式",
    "评分",
    "评价数",
    "供应商/向导名",
    "抓取时间",
    "中文品类",
    "备注",
]


def norm(s: str | None) -> str:
    return (s or "").replace("\xa0", " ").replace("🧭 ", "").strip()


def duration_from(start, end) -> str:
    if not start or not end:
        return ""
    try:
        s = datetime.fromisoformat(start)
        e = datetime.fromisoformat(end)
        mins = int((e - s).total_seconds() // 60)
        if mins <= 0:
            return ""
        if mins % 60 == 0:
            return f"{mins // 60}小时"
        if mins >= 60:
            return f"{mins // 60}小时{mins % 60}分"
        return f"{mins}分钟"
    except Exception:
        return ""


def first_offer(obj) -> dict:
    if not obj:
        return {}
    offers = obj.get("offers") or {}
    if isinstance(offers, list):
        return offers[0] if offers else {}
    return offers if isinstance(offers, dict) else {}


def parse_product(rec: dict) -> dict:
    data = rec.get("data") or {}
    prod = data.get("product") or {}
    events = data.get("events") or []
    ev = None
    for e in events:
        if rec.get("id") and rec["id"] in (e.get("url") or ""):
            ev = e
            break
    if ev is None and events:
        ev = events[0]
    ev = ev or {}
    title = norm(prod.get("name") or ev.get("name") or data.get("title"))
    desc = norm(prod.get("description") or "")
    ev_desc = norm(ev.get("description") or "")
    ev_desc = re.sub(
        r"Этот билет вы\s*можете оплатить картой[^|]{0,180}",
        "",
        ev_desc,
        flags=re.I,
    ).strip()
    blob = " | ".join(x for x in [title, desc, ev_desc] if x)
    offers = first_offer(prod) or first_offer(ev)
    price = offers.get("price")
    if isinstance(price, list):
        price = price[0] if price else None
    try:
        price_f = float(price) if price not in (None, "") else None
    except Exception:
        price_f = None
    currency = offers.get("priceCurrency") or ""
    rating = (prod.get("aggregateRating") or {})
    loc = ev.get("location") or {}
    if isinstance(loc, list):
        loc = loc[0] if loc else {}
    addr = loc.get("address") or []
    if isinstance(addr, dict):
        addr = [addr]
    city_ru = ""
    country = ""
    if addr:
        city_ru = addr[0].get("addressLocality") or ""
        ctry = addr[0].get("addressCountry")
        if isinstance(ctry, list):
            ctry = ctry[0] if ctry else {}
        country = ctry.get("name") if isinstance(ctry, dict) else (ctry or "")
    org = ev.get("organizer") or {}
    kinds = rec.get("source_kinds") or []
    pages = rec.get("source_pages") or []
    listing = rec.get("listing_schema") or {}
    if not title:
        title = norm(listing.get("name"))
    if not ev_desc and listing.get("description"):
        ev_desc = norm(listing.get("description"))
        blob = " | ".join(x for x in [title, desc, ev_desc] if x)
    city_hint = rec.get("city_hint") or loc.get("name") or ""
    return {
        "id": rec.get("id"),
        "url": rec.get("url") or f"https://experience.tripster.ru/experience/{rec.get('id')}/",
        "ok": rec.get("ok"),
        "title": title,
        "desc": desc,
        "ev_desc": ev_desc,
        "blob": blob,
        "price": price_f,
        "currency": currency,
        "rating": rating.get("ratingValue"),
        "reviews": rating.get("reviewCount"),
        "guide": org.get("name") or "",
        "city_en": loc.get("name") or city_hint,
        "city_ru": city_ru,
        "country": country,
        "duration": duration_from(ev.get("startDate"), ev.get("endDate"))
        or duration_from(listing.get("startDate"), listing.get("endDate")),
        "source_kinds": kinds,
        "source_pages": pages,
        "city_hint": city_hint,
        "fetched_at": (rec.get("fetched_at") or "")[:19].replace("T", " "),
        "error": rec.get("error"),
    }


def in_scope(p: dict) -> tuple[bool, str]:
    """Keep only transfer/ticket-ish items. Drop generic city-page tours."""
    kinds = set(p["source_kinds"])
    title = p["title"]
    blob = p["blob"]
    has_tf = bool(TRANSFER_RE.search(title) or TRANSFER_RE.search(blob))
    has_tk = bool(TICKET_RE.search(title) or TICKET_RE.search(blob))
    on_tf_page = bool(kinds & {"transfer", "airport_transfer"})
    on_tk_page = "ticket" in kinds
    only_broad = kinds <= {"city"} or kinds <= {"city", "airport_transfer"} and "iz-aeroporta" in " ".join(
        p["source_pages"]
    )
    # Always keep if listed on dedicated transfer/ticket pages
    if on_tf_page or on_tk_page:
        return True, "listed"
    if has_tf or has_tk:
        return True, "keyword"
    if only_broad:
        return False, "broad_listing"
    return False, "unrelated"


def classify(p: dict) -> dict:
    title = p["title"]
    own = " | ".join(x for x in [p["title"], p["desc"], p.get("ev_desc") or ""] if x)
    blob = p["blob"]
    kinds = set(p["source_kinds"])
    has_tf_title = bool(TRANSFER_RE.search(title))
    has_tk_title = bool(TICKET_RE.search(title))
    has_tf = has_tf_title or bool(TRANSFER_RE.search(own))
    has_tk = has_tk_title or bool(TICKET_RE.search(own))
    has_guide = bool(GUIDE_RE.search(own) or GUIDE_RE.search(title))
    has_sight = bool(SIGHTSEEING_RE.search(own))
    tourish = bool(TOUR_EXCLUDE_RE.search(title) or TOUR_EXCLUDE_RE.search(p["desc"] or ""))

    plus_ticket = bool(
        re.search(
            r"трансфер.+\+ ?билет|\+ ?билет|билет, трансфер|трансфер и безлимит|"
            r"комфортный трансфер.+\+ ?безлимитн|билеты включ",
            own,
            re.I,
        )
    ) or (has_tf_title and has_tk_title)

    # 1) 票加车 first
    if plus_ticket:
        return {
            "cat": "transfer_plus_ticket",
            "reason": f"标题/描述同时出现接送和门票。摘录：「{clip(title)}」。",
            "zh": "票加车",
            "has_tf": "是",
            "has_tk": "是",
            "has_gd": "待确认" if has_guide else "否",
            "meet": "上门接",
            "unit": "每单",
            "note": "接送+门票套餐",
        }

    # 2) transfer with guiding / sightseeing -> not pure
    if has_tf and has_guide:
        return {
            "cat": "uncertain",
            "reason": "含接送，但文案有车上讲解/导览（纯接送不收导览内容）。" + quote_hit(blob, GUIDE_RE),
            "zh": "待确认",
            "has_tf": "是",
            "has_tk": "否" if not has_tk else "待确认",
            "has_gd": "是",
            "meet": "上门接",
            "unit": "每车",
            "note": "车上讲解，按规则不进纯接送",
        }

    if has_tf and has_sight and not has_tk_title:
        return {
            "cat": "uncertain",
            "reason": "标题/描述是接送，但内容是行程/游览，不是点对点。" + quote_hit(blob, SIGHTSEEING_RE),
            "zh": "待确认",
            "has_tf": "是",
            "has_tk": "待确认",
            "has_gd": "待确认",
            "meet": "上门接",
            "unit": "每单",
            "note": "接送+游览，门票是否包含未写清",
        }

    # 3) pure transfer
    if has_tf_title or (has_tf and "transfer" in kinds):
        if tourish and not has_tf_title:
            return uncertain_row(p, "标题更像讲解团，虽出现在接送相关页。")
        point = bool(
            re.search(
                r"аэропорт|отел|адрес|точки А в точку Б|из города в аэропорт|"
                r"между аэропорт|до отеля|по нужному адресу|указанному адресу|"
                r"airport|hotel",
                blob,
                re.I,
            )
        )
        if point or has_tf_title:
            return {
                "cat": "transfer_only",
                "reason": "点对点运输，未见门票套餐，未见导览/行程。" + quote_hit(blob, TRANSFER_RE),
                "zh": "纯接送",
                "has_tf": "是",
                "has_tk": "否",
                "has_gd": "否",
                "meet": "上门接",
                "unit": "每车",
                "note": "",
            }

    # 4) ticket only
    city_tour_title = bool(
        re.search(
            r"экскурси|за 1 день|за один день|многое другое|знакомств|"
            r"в мини-группе|обзорн",
            title,
            re.I,
        )
    )
    admission_like = (
        bool(ADMISSION_TITLE_RE.search(title))
        and "ticket" in kinds
        and not city_tour_title
        and not tourish
    )
    if has_tk_title or (has_tk and "ticket" in kinds) or admission_like:
        if tourish and not has_tk_title:
            return uncertain_row(p, "出现在门票标签，但内容是沙丘游/越野/讲解团，不是入场券。")
        if has_tf and not plus_ticket:
            return {
                "cat": "uncertain",
                "reason": "有门票特征，同时出现接送，但未写成明确套餐。",
                "zh": "待确认",
                "has_tf": "待确认",
                "has_tk": "是",
                "has_gd": "是" if has_guide else "否",
                "meet": "待确认",
                "unit": "每人",
                "note": "",
            }
        if has_guide and re.search(r"аудиоэкскурси|с гидом|обзорная экскурси", own, re.I):
            return {
                "cat": "uncertain",
                "reason": "门票+讲解/音频城市游，不是单门票。" + quote_hit(own, GUIDE_RE),
                "zh": "待确认",
                "has_tf": "否",
                "has_tk": "是",
                "has_gd": "是",
                "meet": "指定点",
                "unit": "每人",
                "note": "门票+导览",
            }
        if tourish and not has_tk_title:
            return uncertain_row(p, "出现在门票标签，但标题是讲解团。")
        return {
            "cat": "ticket_only",
            "reason": "本质是入场券/门票/演出票，未见导游陪同，未见接送。"
            + (quote_hit(title, TICKET_RE) or quote_hit(title, ADMISSION_TITLE_RE)),
            "zh": "单门票",
            "has_tf": "否",
            "has_tk": "是",
            "has_gd": "否",
            "meet": "指定点",
            "unit": "每人",
            "note": "联票如无导游/接送仍算单门票" if re.search(r"комбо|combo|2 в 1|2 в 1", own, re.I) else "",
        }

    return uncertain_row(p, "无法同时满足纯接送、单门票或票加车。")


def uncertain_row(p: dict, why: str) -> dict:
    return {
        "cat": "uncertain",
        "reason": why + f" 标题：「{clip(p['title'])}」。",
        "zh": "待确认",
        "has_tf": "待确认",
        "has_tk": "待确认",
        "has_gd": "待确认",
        "meet": "待确认",
        "unit": "",
        "note": "",
    }


def clip(s: str, n: int = 90) -> str:
    s = re.sub(r"\s+", " ", s).strip()
    return s if len(s) <= n else s[: n - 1] + "…"


def quote_hit(text: str, rgx: re.Pattern) -> str:
    m = rgx.search(text or "")
    if not m:
        return ""
    start = max(0, m.start() - 18)
    end = min(len(text), m.end() + 36)
    return f" 原文：「{clip(text[start:end], 80)}」"


def cny_of(price, currency):
    if price is None or not currency:
        return None, None
    rate = TO_CNY.get(currency.upper())
    if not rate:
        return None, None
    return round(price * rate, 2), rate


def city_label(p: dict) -> str:
    hint = p.get("city_hint") or p.get("city_en") or ""
    ru = p.get("city_ru") or ""
    cn = CITY_CN.get(hint, "")
    parts = [x for x in [ru or hint, hint if hint and hint != ru else "", cn] if x]
    # unique preserve
    seen = []
    for x in parts:
        if x not in seen:
            seen.append(x)
    return " / ".join(seen) if seen else hint


def country_label(p: dict) -> str:
    hint = p.get("city_hint") or ""
    mapped = CITY_COUNTRY.get(hint)
    if mapped:
        return mapped
    return p.get("country") or ""


def tag_label(p: dict) -> str:
    kinds = p["source_kinds"]
    mp = {
        "transfer": "Трансферы",
        "airport_transfer": "Трансферы в/из аэропорта",
        "ticket": "Билеты",
        "city": "Городская витрина",
    }
    return " / ".join(mp.get(k, k) for k in kinds)


def to_row(p: dict, cls: dict) -> list:
    cny, rate = cny_of(p["price"], p["currency"])
    return [
        p["id"],
        p["url"],
        p["title"],
        city_label(p),
        country_label(p),
        cls["cat"],
        cls["reason"],
        tag_label(p),
        p["duration"],
        p["price"],
        p["currency"],
        cny,
        rate,
        FX_DATE if rate else "",
        cls["unit"],
        cls["has_tf"],
        cls["has_tk"],
        cls["has_gd"],
        cls["meet"],
        float(p["rating"]) if p.get("rating") not in (None, "") else None,
        int(p["reviews"]) if p.get("reviews") not in (None, "") else None,
        p.get("guide") or "",
        p.get("fetched_at") or "",
        cls["zh"],
        cls["note"],
    ]


header_fill = PatternFill("solid", fgColor="1F4E79")
header_font = Font(name="PingFang SC", bold=True, color="FFFFFF", size=11)
body_font = Font(name="PingFang SC", size=10)
thin = Border(
    left=Side(style="thin", color="D0D7DE"),
    right=Side(style="thin", color="D0D7DE"),
    top=Side(style="thin", color="D0D7DE"),
    bottom=Side(style="thin", color="D0D7DE"),
)
CAT_FILL = {
    "transfer_only": PatternFill("solid", fgColor="E8F5E9"),
    "ticket_only": PatternFill("solid", fgColor="E3F2FD"),
    "transfer_plus_ticket": PatternFill("solid", fgColor="FFF3E0"),
    "uncertain": PatternFill("solid", fgColor="FFF8E1"),
}
WIDTHS = {
    "A": 12,
    "B": 50,
    "C": 46,
    "D": 22,
    "E": 20,
    "F": 20,
    "G": 54,
    "H": 32,
    "I": 12,
    "J": 10,
    "K": 8,
    "L": 14,
    "M": 10,
    "N": 12,
    "O": 10,
    "P": 12,
    "Q": 12,
    "R": 12,
    "S": 12,
    "T": 8,
    "U": 10,
    "V": 16,
    "W": 20,
    "X": 12,
    "Y": 28,
}


def write_sheet(ws, name, rows, cat):
    ws.title = name
    ws.freeze_panes = "A2"
    for i, h in enumerate(COLS, 1):
        cell = ws.cell(1, i, h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin
    ws.row_dimensions[1].height = 28
    for ri, row in enumerate(rows, 2):
        for ci, val in enumerate(row, 1):
            cell = ws.cell(ri, ci, val)
            cell.font = body_font
            cell.border = thin
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if ci == 6:
                cell.fill = CAT_FILL[cat]
                cell.alignment = Alignment(horizontal="center", vertical="center")
            if ci in (10, 12, 13, 20):
                cell.number_format = "0.00"
        ws.row_dimensions[ri].height = 46
    last = max(2, len(rows) + 1)
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLS))}{last}"
    for col, w in WIDTHS.items():
        ws.column_dimensions[col].width = w


def main():
    products = json.loads(SRC_PRODUCTS.read_text())
    listings = json.loads(SRC_LISTINGS.read_text()) if SRC_LISTINGS.exists() else {}
    fails = json.loads(SRC_FAILS.read_text()) if SRC_FAILS.exists() else []

    parsed = []
    skipped = []
    failed_urls = []
    for rec in products.values():
        p = parse_product(rec)
        if not rec.get("ok"):
            failed_urls.append(rec.get("url") or p["url"])
            continue
        keep, why = in_scope(p)
        if not keep:
            skipped.append({**p, "skip_reason": why})
            continue
        cls = classify(p)
        parsed.append((p, cls))

    buckets = {
        "transfer_only": [],
        "ticket_only": [],
        "transfer_plus_ticket": [],
        "uncertain": [],
    }
    for p, cls in parsed:
        buckets[cls["cat"]].append((p, cls))

    for cat in buckets:
        buckets[cat].sort(key=lambda x: (x[0].get("city_hint") or "", x[0].get("title") or ""))

    wb = Workbook()
    write_sheet(wb.active, "transfer_only", [to_row(p, c) for p, c in buckets["transfer_only"]], "transfer_only")
    write_sheet(
        wb.create_sheet(),
        "ticket_only",
        [to_row(p, c) for p, c in buckets["ticket_only"]],
        "ticket_only",
    )
    write_sheet(
        wb.create_sheet(),
        "transfer_plus_ticket",
        [to_row(p, c) for p, c in buckets["transfer_plus_ticket"]],
        "transfer_plus_ticket",
    )
    write_sheet(
        wb.create_sheet(),
        "uncertain",
        [to_row(p, c) for p, c in buckets["uncertain"]],
        "uncertain",
    )

    # stats
    st = wb.create_sheet("stats")
    st["A1"] = "Tripster 纯接送 / 单门票 / 票加车 抓取统计"
    st["A1"].font = Font(name="PingFang SC", bold=True, size=14, color="1F4E79")
    st.merge_cells("A1:G1")
    meta = [
        ["抓取日期", "2026-08-17"],
        ["城市", "Dubai, Istanbul, Bangkok, Phuket, Bali"],
        ["站点", "https://experience.tripster.ru/"],
        ["robots.txt", "遵守：未请求 /api，未使用带 ? 的 URL；只打开公开列表/详情"],
        ["汇率", FX_NOTE],
        ["USD→CNY", USD_CNY],
        ["规则变更", "车上讲解不进纯接送；送到门口算纯接送；联票无导游/接送算单门票；票加车单独分类"],
        ["详情成功", sum(1 for r in products.values() if r.get("ok"))],
        ["详情失败", len(failed_urls)],
        ["列表页成功", sum(1 for r in listings.values() if r.get("ok"))],
        ["因城市页普通团丢弃", len(skipped)],
        ["入表合计", sum(len(v) for v in buckets.values())],
    ]
    for i, (k, v) in enumerate(meta, 3):
        st.cell(i, 1, k).font = Font(name="PingFang SC", bold=True, size=10)
        st.cell(i, 2, v).font = Font(name="PingFang SC", size=10)
        st.merge_cells(start_row=i, start_column=2, end_row=i, end_column=6)

    st["A17"] = "分城市汇总"
    st["A17"].font = Font(name="PingFang SC", bold=True, size=12, color="1F4E79")
    heads = ["城市", "transfer_only", "ticket_only", "transfer_plus_ticket", "uncertain", "合计", "失败URL"]
    for i, h in enumerate(heads, 1):
        c = st.cell(18, i, h)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center")

    cities = ["Dubai", "Istanbul", "Bangkok", "Phuket", "Bali"]
    fail_by_city = {c: 0 for c in cities}
    for url in failed_urls:
        for c in cities:
            if f"/{c}/" in url or any(
                (products.get(extract_id(url)) or {}).get("city_hint") == c for _ in [0]
            ):
                fail_by_city[c] += 1
                break

    def city_of_row(p):
        return p.get("city_hint") or p.get("city_en") or ""

    for ri, city in enumerate(cities, 19):
        nums = []
        for cat in ["transfer_only", "ticket_only", "transfer_plus_ticket", "uncertain"]:
            nums.append(sum(1 for p, _ in buckets[cat] if city_of_row(p) == city))
        total = sum(nums)
        vals = [city, *nums, total, fail_by_city.get(city, 0)]
        for ci, v in enumerate(vals, 1):
            cell = st.cell(ri, ci, v)
            cell.font = body_font
            cell.alignment = Alignment(horizontal="center")
            if ci == 2:
                cell.fill = CAT_FILL["transfer_only"]
            if ci == 3:
                cell.fill = CAT_FILL["ticket_only"]
            if ci == 4:
                cell.fill = CAT_FILL["transfer_plus_ticket"]
            if ci == 5:
                cell.fill = CAT_FILL["uncertain"]

    st["A26"] = "失败 URL"
    st["A26"].font = Font(name="PingFang SC", bold=True, size=12, color="1F4E79")
    if failed_urls:
        for i, u in enumerate(failed_urls, 27):
            st.cell(i, 1, u).font = body_font
            st.merge_cells(start_row=i, start_column=1, end_row=i, end_column=6)
        fail_end = 26 + len(failed_urls)
    else:
        st["A27"] = "无"
        fail_end = 27

    st.cell(fail_end + 2, 1, "列表页").font = Font(name="PingFang SC", bold=True, size=12, color="1F4E79")
    r = fail_end + 3
    for i, h in enumerate(["城市", "类型", "URL", "卡片数", "状态"], 1):
        c = st.cell(r, i, h)
        c.fill = header_fill
        c.font = header_font
    r += 1
    for rec in listings.values():
        nitems = len((rec.get("data") or {}).get("items") or [])
        st.cell(r, 1, rec.get("city")).font = body_font
        st.cell(r, 2, rec.get("kind")).font = body_font
        st.cell(r, 3, rec.get("url")).font = body_font
        st.cell(r, 4, nitems).font = body_font
        st.cell(r, 5, "ok" if rec.get("ok") else rec.get("error")).font = body_font
        r += 1

    st.column_dimensions["A"].width = 24
    st.column_dimensions["B"].width = 22
    st.column_dimensions["C"].width = 70
    st.column_dimensions["D"].width = 22
    st.column_dimensions["E"].width = 22
    st.column_dimensions["F"].width = 10
    st.column_dimensions["G"].width = 12

    # skipped inventory
    sk = wb.create_sheet("skipped_city_tours")
    sk["A1"] = "从城市总览页抓到、但按规则丢弃的普通讲解团（未入三类表）"
    sk["A1"].font = Font(name="PingFang SC", bold=True, size=12, color="1F4E79")
    for i, h in enumerate(["产品ID", "城市", "标题", "URL"], 1):
        c = sk.cell(2, i, h)
        c.fill = header_fill
        c.font = header_font
    for i, p in enumerate(sorted(skipped, key=lambda x: (x.get("city_hint") or "", x.get("title") or "")), 3):
        sk.cell(i, 1, p["id"]).font = body_font
        sk.cell(i, 2, p.get("city_hint")).font = body_font
        sk.cell(i, 3, p.get("title")).font = body_font
        sk.cell(i, 4, p.get("url")).font = body_font
    for col, w in zip("ABCD", [12, 14, 70, 52]):
        sk.column_dimensions[col].width = w

    OUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_XLSX)
    print("saved", OUT_XLSX, "bytes", OUT_XLSX.stat().st_size)
    print("counts", {k: len(v) for k, v in buckets.items()})
    print("skipped", len(skipped), "failed", len(failed_urls))
    for city in cities:
        print(
            city,
            {cat: sum(1 for p, _ in buckets[cat] if city_of_row(p) == city) for cat in buckets},
        )


def extract_id(url: str) -> str:
    m = re.search(r"/experience/(\d+)/", url or "")
    return m.group(1) if m else ""


if __name__ == "__main__":
    main()
