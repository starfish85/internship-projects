#!/usr/bin/env python3
"""Turn a snapshot.json into the two dated business workbooks."""

from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.hyperlink import Hyperlink

from paths import project_root, tripadvisor_xlsx_path, viator_xlsx_path

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
BODY_FONT = Font(name="Calibri", size=11)
THIN = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)
WRAP = Alignment(wrap_text=True, vertical="center")
FILL_PASS = PatternFill("solid", fgColor="C6EFCE")
FILL_FAIL = PatternFill("solid", fgColor="FFC7CE")
FILL_PENDING = PatternFill("solid", fgColor="FFEB9C")
FILL_ZEBRA = PatternFill("solid", fgColor="F2F2F2")

VIATOR_COLUMNS = [
    ("采集时间", 18),
    ("产品名称", 42),
    ("产品code", 18),
    ("Viator状态", 14),
    ("Viator产品评分", 16),
    ("Viator评价数", 14),
    ("Viator产品链接", 28),
    ("企业详情名称", 28),
    ("企业详情位置", 36),
    ("Tripadvisor透传状态", 20),
    ("Tripadvisor产品链接", 32),
    ("Tripadvisor产品名称", 36),
    ("Tripadvisor产品评分", 18),
    ("Tripadvisor评价数", 16),
    ("Tripadvisor评价类型", 18),
    ("Tripadvisor页面code", 20),
    ("备注", 28),
]

LOCATION_COLUMNS = [
    ("采集时间", 18),
    ("企业位置名称", 32),
    ("企业位置链接", 40),
    ("评分", 10),
    ("评价数", 12),
    ("地理信息", 28),
    ("类别", 16),
    ("展示产品数量", 14),
    ("来源", 32),
]

LOCATION_PRODUCT_COLUMNS = [
    ("采集时间", 18),
    ("企业位置名称", 28),
    ("企业位置链接", 36),
    ("序号", 8),
    ("产品名称", 42),
    ("产品链接", 36),
    ("产品评分", 12),
    ("产品评价数", 12),
    ("产品code", 18),
]


def load_snapshot(path: Path) -> dict[str, Any]:
    data = json.loads(path.read_text(encoding="utf-8"))
    if not data.get("run_id"):
        raise SystemExit(f"snapshot missing run_id: {path}")
    data.setdefault("viator_products", [])
    data.setdefault("enterprise_locations", [])
    return data


def style_header(ws, widths: list[tuple[str, int]]) -> None:
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(widths))}1"
    ws.row_dimensions[1].height = 22
    for idx, (title, width) in enumerate(widths, start=1):
        cell = ws.cell(1, idx, title)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(idx)].width = width


def write_row(ws, row: int, values: list[Any], url_cols: set[int] | None = None) -> None:
    url_cols = url_cols or set()
    for col, value in enumerate(values, start=1):
        cell = ws.cell(row, col, "" if value is None else value)
        cell.font = BODY_FONT
        cell.alignment = WRAP
        cell.border = THIN
        if row % 2 == 0:
            cell.fill = FILL_ZEBRA
        if col in url_cols and value:
            cell.hyperlink = Hyperlink(ref=cell.coordinate, target=str(value), tooltip=str(value))
            cell.font = Font(name="Calibri", size=11, color="0563C1", underline="single")
            cell.style = "Hyperlink"


def syndication_fill(status: str) -> PatternFill | None:
    if status == "已透传":
        return FILL_PASS
    if status == "未透传":
        return FILL_FAIL
    if status == "未核验":
        return FILL_PENDING
    return None


def write_viator(path: Path, snap: dict[str, Any]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "已上线产品"
    style_header(ws, VIATOR_COLUMNS)
    run_id = snap["run_id"]
    url_cols = {7, 11}
    for i, p in enumerate(snap["viator_products"], start=2):
        values = [
            run_id,
            p.get("product_name"),
            p.get("product_code"),
            p.get("viator_status") or "已上线",
            p.get("viator_rating"),
            p.get("viator_review_count"),
            p.get("viator_url"),
            p.get("enterprise_name"),
            p.get("enterprise_location_text"),
            p.get("syndication_status") or "未核验",
            p.get("ta_product_url"),
            p.get("ta_product_name"),
            p.get("ta_product_rating"),
            p.get("ta_product_review_count"),
            p.get("ta_review_type"),
            p.get("ta_product_code_found"),
            p.get("notes"),
        ]
        write_row(ws, i, values, url_cols)
        fill = syndication_fill(str(values[9] or ""))
        if fill:
            ws.cell(i, 10).fill = fill
    ws.auto_filter.ref = f"A1:{get_column_letter(len(VIATOR_COLUMNS))}{max(1, ws.max_row)}"
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def write_tripadvisor(path: Path, snap: dict[str, Any]) -> None:
    wb = Workbook()
    loc_ws = wb.active
    loc_ws.title = "企业位置"
    style_header(loc_ws, LOCATION_COLUMNS)
    prod_ws = wb.create_sheet("位置下产品")
    style_header(prod_ws, LOCATION_PRODUCT_COLUMNS)

    run_id = snap["run_id"]
    prod_row = 2
    for i, loc in enumerate(snap["enterprise_locations"], start=2):
        products = loc.get("products") or []
        count = loc.get("product_count")
        if count is None:
            count = len(products)
        sources = loc.get("sources") or []
        if isinstance(sources, list):
            sources = "、".join(str(s) for s in sources)
        write_row(
            loc_ws,
            i,
            [
                run_id,
                loc.get("name"),
                loc.get("url"),
                loc.get("rating"),
                loc.get("review_count"),
                loc.get("geo"),
                loc.get("category"),
                count,
                sources,
            ],
            url_cols={3},
        )
        for prod in products:
            write_row(
                prod_ws,
                prod_row,
                [
                    run_id,
                    loc.get("name"),
                    loc.get("url"),
                    prod.get("position"),
                    prod.get("name"),
                    prod.get("url"),
                    prod.get("rating"),
                    prod.get("review_count"),
                    prod.get("product_code"),
                ],
                url_cols={3, 6},
            )
            prod_row += 1

    loc_ws.auto_filter.ref = f"A1:{get_column_letter(len(LOCATION_COLUMNS))}{max(1, loc_ws.max_row)}"
    prod_ws.auto_filter.ref = (
        f"A1:{get_column_letter(len(LOCATION_PRODUCT_COLUMNS))}{max(1, prod_ws.max_row)}"
    )
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Write Viator and Tripadvisor xlsx from a snapshot")
    parser.add_argument("--snapshot", required=True, help="Path to snapshot.json")
    parser.add_argument("--project", default=None, help="Project root (optional)")
    args = parser.parse_args(argv)

    snap_file = Path(args.snapshot).expanduser().resolve()
    if not snap_file.is_file():
        raise SystemExit(f"snapshot not found: {snap_file}")

    snap = load_snapshot(snap_file)
    root = project_root(args.project)
    v_path = viator_xlsx_path(root, snap["run_id"])
    t_path = tripadvisor_xlsx_path(root, snap["run_id"])
    write_viator(v_path, snap)
    write_tripadvisor(t_path, snap)
    print(v_path)
    print(t_path)
    return 0


if __name__ == "__main__":
    sys.exit(main())
