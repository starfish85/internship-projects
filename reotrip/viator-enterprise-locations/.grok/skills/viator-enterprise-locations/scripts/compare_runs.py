#!/usr/bin/env python3
"""Date-to-date comparison of two snapshots. Same calendar day keeps the latest run."""

from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from paths import comparison_xlsx_path, project_root

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
BODY_FONT = Font(name="Calibri", size=11)
THIN = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)
WRAP = Alignment(wrap_text=True, vertical="center")
FILL_NEW = PatternFill("solid", fgColor="C6EFCE")
FILL_GONE = PatternFill("solid", fgColor="FFC7CE")
FILL_CHANGED = PatternFill("solid", fgColor="FFEB9C")
FILL_UP = PatternFill("solid", fgColor="C6EFCE")
FILL_DOWN = PatternFill("solid", fgColor="FFC7CE")
FILL_ZEBRA = PatternFill("solid", fgColor="F2F2F2")


def load_json(path: Path) -> dict[str, Any]:
    return json.loads(path.read_text(encoding="utf-8"))


def list_snapshots(root: Path) -> list[tuple[str, Path, dict[str, Any]]]:
    raw = root / "data" / "raw"
    found: list[tuple[str, Path, dict[str, Any]]] = []
    if not raw.is_dir():
        return found
    for snap in sorted(raw.glob("*/snapshot.json")):
        data = load_json(snap)
        run_id = data.get("run_id") or snap.parent.name
        found.append((run_id, snap, data))
    found.sort(key=lambda x: x[0])
    return found


def date_of(run_id: str) -> str:
    return run_id[:8]


def latest_per_date(snaps: list[tuple[str, Path, dict[str, Any]]]) -> dict[str, tuple[str, Path, dict[str, Any]]]:
    by_date: dict[str, tuple[str, Path, dict[str, Any]]] = {}
    for item in snaps:
        by_date[date_of(item[0])] = item
    return by_date


def pick_pair(
    snaps: list[tuple[str, Path, dict[str, Any]]],
    from_date: str | None,
    to_date: str | None,
) -> tuple[tuple[str, Path, dict[str, Any]], tuple[str, Path, dict[str, Any]]]:
    by_date = latest_per_date(snaps)
    dates = sorted(by_date)
    if len(dates) < 2 and not (from_date and to_date):
        raise SystemExit("Need at least two collection dates to compare.")
    if to_date and from_date:
        if from_date not in by_date:
            raise SystemExit(f"No snapshot for date {from_date}")
        if to_date not in by_date:
            raise SystemExit(f"No snapshot for date {to_date}")
        return by_date[from_date], by_date[to_date]
    if to_date:
        if to_date not in by_date:
            raise SystemExit(f"No snapshot for date {to_date}")
        earlier = [d for d in dates if d < to_date]
        if not earlier:
            raise SystemExit(f"No earlier date than {to_date} to compare against.")
        return by_date[earlier[-1]], by_date[to_date]
    return by_date[dates[-2]], by_date[dates[-1]]


def num(value: Any) -> float | None:
    if value is None or value == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def fmt_delta(old: Any, new: Any) -> tuple[str, float | None]:
    a, b = num(old), num(new)
    if a is None or b is None:
        if old == new:
            return "", None
        return "有变动", None
    delta = b - a
    if delta == 0:
        return "0", 0.0
    sign = "+" if delta > 0 else ""
    if float(delta).is_integer() and float(a).is_integer() and float(b).is_integer():
        return f"{sign}{int(delta)}", delta
    return f"{sign}{delta:.2f}", delta


def loc_key(loc: dict[str, Any]) -> str:
    url = (loc.get("url") or "").strip()
    if url:
        return f"url:{url}"
    name = (loc.get("name") or "").strip()
    geo = (loc.get("geo") or "").strip()
    return f"name:{name}|{geo}"


def prod_under_loc_key(loc: dict[str, Any], prod: dict[str, Any]) -> str:
    p_url = (prod.get("url") or "").strip()
    if p_url:
        return p_url
    code = (prod.get("product_code") or "").strip()
    if code:
        return f"code:{code}"
    return f"name:{(prod.get('name') or '').strip()}"


def style_header(ws, titles: list[str], widths: list[int]) -> None:
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 22
    for idx, title in enumerate(titles, start=1):
        cell = ws.cell(1, idx, title)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(idx)].width = widths[idx - 1]


def paint(cell, fill: PatternFill | None) -> None:
    if fill:
        cell.fill = fill


def append_row(ws, values: list[Any], fill: PatternFill | None = None) -> None:
    row = ws.max_row + 1
    for col, value in enumerate(values, start=1):
        cell = ws.cell(row, col, "" if value is None else value)
        cell.font = BODY_FONT
        cell.alignment = WRAP
        cell.border = THIN
        if fill:
            cell.fill = fill
        elif row % 2 == 0:
            cell.fill = FILL_ZEBRA


def finish_sheet(ws, ncols: int) -> None:
    ws.auto_filter.ref = f"A1:{get_column_letter(ncols)}{max(1, ws.max_row)}"


def build_workbook(
    older: dict[str, Any],
    newer: dict[str, Any],
    from_date: str,
    to_date: str,
) -> Workbook:
    old_products = {p.get("product_code"): p for p in older.get("viator_products") or [] if p.get("product_code")}
    new_products = {p.get("product_code"): p for p in newer.get("viator_products") or [] if p.get("product_code")}
    old_locs = {loc_key(l): l for l in older.get("enterprise_locations") or []}
    new_locs = {loc_key(l): l for l in newer.get("enterprise_locations") or []}

    added_codes = [c for c in new_products if c not in old_products]
    removed_codes = [c for c in old_products if c not in new_products]
    synd_changes = []
    rating_changes = []
    for code in sorted(set(old_products) & set(new_products)):
        a, b = old_products[code], new_products[code]
        old_s = a.get("syndication_status") or ""
        new_s = b.get("syndication_status") or ""
        if old_s != new_s:
            synd_changes.append((code, a, b, old_s, new_s))
        label, delta = fmt_delta(a.get("ta_product_rating"), b.get("ta_product_rating"))
        if delta not in (None, 0.0) or (label == "有变动"):
            rating_changes.append(("Tripadvisor产品评分", code, a, b, a.get("ta_product_rating"), b.get("ta_product_rating"), label))
        label, delta = fmt_delta(a.get("viator_rating"), b.get("viator_rating"))
        if delta not in (None, 0.0) or (label == "有变动"):
            rating_changes.append(("Viator产品评分", code, a, b, a.get("viator_rating"), b.get("viator_rating"), label))

    added_locs = [k for k in new_locs if k not in old_locs]
    removed_locs = [k for k in old_locs if k not in new_locs]
    loc_rating_changes = []
    loc_product_churn = []
    loc_listed_rating_changes = []
    for key in sorted(set(old_locs) & set(new_locs)):
        a, b = old_locs[key], new_locs[key]
        label, delta = fmt_delta(a.get("rating"), b.get("rating"))
        count_label, _ = fmt_delta(a.get("product_count", len(a.get("products") or [])), b.get("product_count", len(b.get("products") or [])))
        review_label, _ = fmt_delta(a.get("review_count"), b.get("review_count"))
        if delta not in (None, 0.0) or label == "有变动" or count_label not in ("", "0") or review_label not in ("", "0"):
            loc_rating_changes.append((a, b, label, count_label, review_label))
        old_listed = {prod_under_loc_key(a, p): p for p in a.get("products") or []}
        new_listed = {prod_under_loc_key(b, p): p for p in b.get("products") or []}
        for pk in new_listed:
            if pk not in old_listed:
                loc_product_churn.append(("新增展示", b, new_listed[pk]))
        for pk in old_listed:
            if pk not in new_listed:
                loc_product_churn.append(("不再展示", a, old_listed[pk]))
        for pk in set(old_listed) & set(new_listed):
            op, np_ = old_listed[pk], new_listed[pk]
            rlabel, rdelta = fmt_delta(op.get("rating"), np_.get("rating"))
            if rdelta not in (None, 0.0) or rlabel == "有变动":
                loc_listed_rating_changes.append((b, op, np_, rlabel))

    wb = Workbook()

    overview = wb.active
    overview.title = "总览"
    style_header(overview, ["项目", "旧日期", "新日期", "说明"], [28, 22, 22, 48])
    def n_synd(products: dict, status: str) -> int:
        return sum(1 for p in products.values() if p.get("syndication_status") == status)

    rows = [
        ("对比日期", from_date, to_date, f"{older.get('run_id')} → {newer.get('run_id')}"),
        ("Viator产品数", len(old_products), len(new_products), f"新增 {len(added_codes)} / 消失 {len(removed_codes)}"),
        ("已透传", n_synd(old_products, "已透传"), n_synd(new_products, "已透传"), f"透传状态变动 {len(synd_changes)}"),
        ("未透传", n_synd(old_products, "未透传"), n_synd(new_products, "未透传"), ""),
        ("未核验", n_synd(old_products, "未核验"), n_synd(new_products, "未核验"), ""),
        ("企业位置数", len(old_locs), len(new_locs), f"新增 {len(added_locs)} / 消失 {len(removed_locs)}"),
        ("位置下产品条数", sum(len(l.get("products") or []) for l in old_locs.values()), sum(len(l.get("products") or []) for l in new_locs.values()), f"展示增减 {len(loc_product_churn)}"),
        ("产品评分变动", len([r for r in rating_changes]), len(loc_listed_rating_changes), "左=Viator/透传产品评分条数，右=位置下产品评分条数"),
        ("企业位置评分/数量变动", len(loc_rating_changes), "", ""),
    ]
    for r in rows:
        append_row(overview, list(r), FILL_CHANGED if r[0].endswith("变动") or "变动" in str(r[3]) else None)
    finish_sheet(overview, 4)

    ws = wb.create_sheet("透传状态变动")
    cols = ["变动", "产品code", "产品名称", "旧透传状态", "新透传状态", "旧TA链接", "新TA链接", "旧TA评分", "新TA评分"]
    style_header(ws, cols, [14, 18, 36, 16, 16, 32, 32, 12, 12])
    for code, a, b, old_s, new_s in synd_changes:
        fill = FILL_CHANGED
        if old_s != "已透传" and new_s == "已透传":
            fill = FILL_NEW
        elif old_s == "已透传" and new_s != "已透传":
            fill = FILL_GONE
        append_row(ws, ["透传翻转", code, b.get("product_name") or a.get("product_name"), old_s, new_s, a.get("ta_product_url"), b.get("ta_product_url"), a.get("ta_product_rating"), b.get("ta_product_rating")], fill)
    finish_sheet(ws, len(cols))

    ws = wb.create_sheet("Viator产品增减")
    cols = ["变动", "产品code", "产品名称", "企业详情名称", "透传状态", "Viator评分"]
    style_header(ws, cols, [12, 18, 42, 28, 14, 12])
    for code in sorted(added_codes):
        p = new_products[code]
        append_row(ws, ["新增上线", code, p.get("product_name"), p.get("enterprise_name"), p.get("syndication_status"), p.get("viator_rating")], FILL_NEW)
    for code in sorted(removed_codes):
        p = old_products[code]
        append_row(ws, ["本次消失", code, p.get("product_name"), p.get("enterprise_name"), p.get("syndication_status"), p.get("viator_rating")], FILL_GONE)
    finish_sheet(ws, len(cols))

    ws = wb.create_sheet("企业位置增减")
    cols = ["变动", "企业位置名称", "链接", "评分", "评价数", "展示产品数量", "地理信息"]
    style_header(ws, cols, [12, 32, 40, 10, 10, 14, 28])
    for k in sorted(added_locs):
        l = new_locs[k]
        append_row(ws, ["新增位置", l.get("name"), l.get("url"), l.get("rating"), l.get("review_count"), l.get("product_count", len(l.get("products") or [])), l.get("geo")], FILL_NEW)
    for k in sorted(removed_locs):
        l = old_locs[k]
        append_row(ws, ["位置消失", l.get("name"), l.get("url"), l.get("rating"), l.get("review_count"), l.get("product_count", len(l.get("products") or [])), l.get("geo")], FILL_GONE)
    finish_sheet(ws, len(cols))

    ws = wb.create_sheet("企业位置评分变动")
    cols = ["企业位置名称", "链接", "旧评分", "新评分", "评分差", "旧评价数", "新评价数", "旧产品数", "新产品数", "产品数差"]
    style_header(ws, cols, [28, 40, 10, 10, 10, 12, 12, 12, 12, 12])
    for a, b, rlabel, clabel, _review in loc_rating_changes:
        fill = FILL_CHANGED
        _, delta = fmt_delta(a.get("rating"), b.get("rating"))
        if delta is not None and delta > 0:
            fill = FILL_UP
        elif delta is not None and delta < 0:
            fill = FILL_DOWN
        append_row(
            ws,
            [
                b.get("name") or a.get("name"),
                b.get("url") or a.get("url"),
                a.get("rating"),
                b.get("rating"),
                rlabel,
                a.get("review_count"),
                b.get("review_count"),
                a.get("product_count", len(a.get("products") or [])),
                b.get("product_count", len(b.get("products") or [])),
                clabel,
            ],
            fill,
        )
    finish_sheet(ws, len(cols))

    ws = wb.create_sheet("位置下产品增减")
    cols = ["变动", "企业位置名称", "企业位置链接", "产品名称", "产品链接", "产品评分", "产品code"]
    style_header(ws, cols, [12, 28, 36, 42, 36, 12, 18])
    for kind, loc, prod in loc_product_churn:
        fill = FILL_NEW if kind == "新增展示" else FILL_GONE
        append_row(ws, [kind, loc.get("name"), loc.get("url"), prod.get("name"), prod.get("url"), prod.get("rating"), prod.get("product_code")], fill)
    finish_sheet(ws, len(cols))

    ws = wb.create_sheet("产品评分变动")
    cols = ["维度", "标识", "名称", "所属企业位置", "旧评分", "新评分", "差值"]
    style_header(ws, cols, [22, 22, 42, 28, 12, 12, 12])
    for dim, code, a, b, old_r, new_r, label in rating_changes:
        _, delta = fmt_delta(old_r, new_r)
        fill = FILL_UP if (delta or 0) > 0 else FILL_DOWN if (delta or 0) < 0 else FILL_CHANGED
        append_row(ws, [dim, code, b.get("product_name") or a.get("product_name"), b.get("enterprise_name") or a.get("enterprise_name"), old_r, new_r, label], fill)
    for loc, op, np_, rlabel in loc_listed_rating_changes:
        _, delta = fmt_delta(op.get("rating"), np_.get("rating"))
        fill = FILL_UP if (delta or 0) > 0 else FILL_DOWN if (delta or 0) < 0 else FILL_CHANGED
        append_row(ws, ["位置下产品评分", np_.get("product_code") or np_.get("url") or op.get("name"), np_.get("name") or op.get("name"), loc.get("name"), op.get("rating"), np_.get("rating"), rlabel], fill)
    finish_sheet(ws, len(cols))

    return wb


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Compare two dated snapshots")
    parser.add_argument("--project", default=None)
    parser.add_argument("--from-date", default=None, help="YYYYMMDD")
    parser.add_argument("--to-date", default=None, help="YYYYMMDD")
    args = parser.parse_args(argv)

    root = project_root(args.project)
    snaps = list_snapshots(root)
    if not snaps:
        raise SystemExit(f"No snapshots under {root / 'data' / 'raw'}")

    older, newer = pick_pair(snaps, args.from_date, args.to_date)
    from_date, to_date = date_of(older[0]), date_of(newer[0])
    wb = build_workbook(older[2], newer[2], from_date, to_date)
    out = comparison_xlsx_path(root, from_date, to_date)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    print(out)
    print(f"{older[0]} -> {newer[0]}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
