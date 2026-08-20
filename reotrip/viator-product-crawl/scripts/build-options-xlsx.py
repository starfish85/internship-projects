#!/usr/bin/env python3
"""Build 4-sheet workbook: name / code / option (+ ticket extra columns)."""
from __future__ import annotations

import json
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path("/Users/mac/viator产品爬取")
src = json.loads((ROOT / "data/viator-options.json").read_text())
products = src["products"]

# simple cancel mapping — only fill when unambiguous
CANCEL_MAP = {
    "ALL_SALES_FINAL": "不可取消",
    "STANDARD_1_DAY": "出行日期前1天不可取消",
    "STANDARD_24HOURS": "出行日期前1天不可取消",
    "STANDARD_2_DAYS": "出行日期前2天不可取消",
    "STANDARD_3_DAYS": "出行日期前3天不可取消",
    "STANDARD_7_DAYS": "出行日期前7天不可取消",
}

YELLOW = PatternFill("solid", fgColor="FFF258")
HEADER = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
BODY = Font(name="Calibri", size=11)
THIN = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)
WRAP = Alignment(vertical="center", wrap_text=True)


def option_label(opt: dict) -> str:
    title = (opt.get("title") or "").strip()
    grade = (opt.get("tourGradeCode") or "").strip()
    return title or grade or opt.get("ref") or ""


def cancel_text(p: dict) -> str:
    c = p.get("cancellation") or {}
    code = c.get("type") or ""
    if code in CANCEL_MAP:
        return CANCEL_MAP[code]
    return ""  # leave empty; fill later after supply-chain pass


def apply_header(ws, headers):
    for col, h in enumerate(headers, 1):
        cell = ws.cell(1, col, h)
        cell.fill = HEADER
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
    ws.row_dimensions[1].height = 22


def paint_row(ws, r, cols, fill):
    for c in range(1, cols + 1):
        cell = ws.cell(r, c)
        cell.font = BODY
        cell.alignment = WRAP
        cell.border = THIN
        cell.fill = fill
    ws.row_dimensions[r].height = 22


def merge_same_product(ws, start_row, end_row):
    """Merge 产品名称 / 产品代码 for consecutive rows of one product (yellow style)."""
    if end_row <= start_row:
        return
    ws.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)
    ws.merge_cells(start_row=start_row, start_column=2, end_row=end_row, end_column=2)
    ws.cell(start_row, 1).alignment = Alignment(vertical="center", wrap_text=True)
    ws.cell(start_row, 2).alignment = Alignment(vertical="center", wrap_text=True)


wb = openpyxl.Workbook()

# --- tickets ---
ticket_headers = ["产品名称", "产品代码", "option", "具体拆分", "门票使用方式", "门票取消规则", "供应商", "供应商链接"]
ws = wb.active
ws.title = "门票产品"
apply_header(ws, ticket_headers)
r = 2
ticket_n = 0
for p in products:
    if p["category"] != "门票产品":
        continue
    opts = [o for o in (p.get("options") or []) if (o.get("status") or "ACTIVE") == "ACTIVE"] or p.get("options") or [{}]
    ticket_n += 1
    start = r
    for opt in opts:
        ws.cell(r, 1, p["title"])
        ws.cell(r, 2, p["productCode"])
        ws.cell(r, 3, option_label(opt))
        ws.cell(r, 6, cancel_text(p))
        paint_row(ws, r, 8, YELLOW)
        r += 1
    merge_same_product(ws, start, r - 1)
for i, w in enumerate([56, 16, 44, 14, 16, 22, 12, 42], 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# --- simple 3-col sheets ---
simple = [
    ("接送产品", "接送产品"),
    ("包车产品", "包车产品"),
    ("日游产品", "日游产品"),
]
counts = {"门票产品": ticket_n}
for sheet_name, cat in simple:
    wsc = wb.create_sheet(sheet_name)
    apply_header(wsc, ["产品名称", "产品代码", "option"])
    rr = 2
    nprod = 0
    for p in products:
        if p["category"] != cat:
            continue
        nprod += 1
        opts = [o for o in (p.get("options") or []) if (o.get("status") or "ACTIVE") == "ACTIVE"] or p.get("options") or [{}]
        start = rr
        for opt in opts:
            wsc.cell(rr, 1, p["title"])
            wsc.cell(rr, 2, p["productCode"])
            wsc.cell(rr, 3, option_label(opt))
            paint_row(wsc, rr, 3, YELLOW)
            rr += 1
        merge_same_product(wsc, start, rr - 1)
    counts[cat] = nprod
    for i, w in enumerate([64, 16, 48], 1):
        wsc.column_dimensions[get_column_letter(i)].width = w

# summary
wss = wb.create_sheet("抓取说明", 0)
wss["A1"] = "Viator 已启用产品 option 抓取"
wss["A1"].font = Font(name="Calibri", size=16, bold=True, color="1F4E79")
wss["A3"] = f"生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M')}"
wss["A4"] = f"产品 {src.get('ok')}/{src.get('count')} 已抓到 option"
wss["A6"] = "分类"
wss["B6"] = "产品数"
wss["A6"].fill = HEADER
wss["B6"].fill = HEADER
wss["A6"].font = HEADER_FONT
wss["B6"].font = HEADER_FONT
for i, cat in enumerate(["门票产品", "接送产品", "包车产品", "日游产品"], 7):
    wss.cell(i, 1, cat)
    wss.cell(i, 2, counts[cat])
wss["A12"] = "门票表目前只填了产品名称 / 代码 / option；取消规则仅在后台政策很明确时预填（如不可取消）。使用方式、拆分、供应商链接下一步再对供应链。"
wss.merge_cells("A12:F12")
wss.column_dimensions["A"].width = 18
wss.column_dimensions["B"].width = 10

out = ROOT / "Viator已启用产品_option.xlsx"
wb.save(out)
print("saved", out)
print("counts", counts)
print("harvest", src.get("ok"), "/", src.get("count"))
