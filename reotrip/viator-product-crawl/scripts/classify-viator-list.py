#!/usr/bin/env python3
"""Merge page0 + pages 1-17, keep ACTIVE only, classify into 4 buckets."""
from __future__ import annotations

import json
import re
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path("/Users/mac/viator产品爬取")
DATA = ROOT / "data"

# --- load all pages ---
by_code = {}
page_files = sorted(DATA.glob("viator-api-page-*.json"), key=lambda p: int(re.search(r"(\d+)", p.stem).group(1)))

# page 0 is not saved yet; we also accept a dedicated file
for pf in page_files:
    raw = json.loads(pf.read_text())
    body = json.loads(raw["text"]) if isinstance(raw.get("text"), str) else raw
    for p in body.get("products") or []:
        by_code[p["productCode"]] = p


def slim(p: dict) -> dict:
    return {
        "productCode": p.get("productCode"),
        "title": (p.get("title") or "").strip(),
        "status": p.get("status"),
        "isActive": p.get("isActive"),
        "connectedOptionCount": p.get("connectedOptionCount"),
        "totalActiveOptionCount": p.get("totalActiveOptionCount"),
        "viatorUrl": p.get("localizedViatorUrl"),
        "quality": p.get("productOverallQualityLevel"),
    }


# --- classification ---
TRANSFER_RE = re.compile(
    r"\b(private transfer|shared transfer|airport transfer|one way transfer|"
    r"transfer to/?from|transfer between|transfer to or from|"
    r"transfers between|transfer to and from|transfers? to from|"
    r"pickup and drop[- ]?off only|point to point)\b",
    re.I,
)
TICKET_WORD_RE = re.compile(
    r"\b(admission tickets?|admission|1-day pass|day pass|adventure pass|"
    r"early entry pass|premier access|e-tickets?|entrance tickets?|"
    r"entry tickets?|observatory tickets?|show tickets?|"
    r"tickets?|day pass|park pass|adventure pass|pass experience|pass)\b",
    re.I,
)
# destination names that look like tickets but may only be drop-off points
DEST_ONLY_RE = re.compile(
    r"\b(disneyland|disneysea|disney resort|ghibli|universal|pop land)\b",
    re.I,
)
DAYTOUR_RE = re.compile(
    r"\b(bus tour|day tour|day trip|hop-?on|sightseeing bus|"
    r"途益|toyo bus|toyobus|group tour|group trip|walking tour)\b",
    re.I,
)
CHARTER_RE = re.compile(
    r"\b(charter|private car|private vehicle|private van|private coach|"
    r"with (a )?private driver|driver only|hire car|包车|"
    r"private hire|chauffeur|alphard|by car|"
    r"driver or optional( english)? guide)\b",
    re.I,
)
PRIVATE_TOUR_RE = re.compile(
    r"\bprivate\b.{0,50}\btour\b|"
    r"\b\d+\s*-?\s*h(our)?s?\b.*\bprivate\b|\bprivate\b.*\b\d+\s*-?\s*h(our)?s?\b",
    re.I,
)
HOURLY_TOUR_RE = re.compile(r"\b\d+\s*-?\s*(hour|h)\b.*\btour\b|\btour\b.*\b\d+\s*-?\s*(hour|h)\b", re.I)
SPA_OR_VENUE_RE = re.compile(r"\b(spa|onsen|mansion|teahouse|theatre|theater| observatory)\b", re.I)


def classify(title: str) -> tuple[str, str, str]:
    """Return (category, confidence, reason). Aligns with task:
    门票 = 单门票 / 门票+接送 / 门票组合
    包车 = 车型+司机 或 车型+司机&导游（Private Tour 先归这里）
    日游 = 目前主要途益 bus tour / group tour，option 多为集合点
    """
    t = title or ""
    has_ticket_word = bool(TICKET_WORD_RE.search(t))
    has_transfer = bool(TRANSFER_RE.search(t) or re.search(r"\bpick-?up\b|\bdrop-?off\b", t, re.I))

    # 纯接送：去迪士尼/乐园，但标题没有 ticket/pass/admission
    if has_transfer and not has_ticket_word and not PRIVATE_TOUR_RE.search(t):
        return "接送产品", "高", "标题是点对点接送（目的地即使是乐园也不含门票）"

    # 门票+接送 / 门票组合
    if has_ticket_word and has_transfer:
        return "门票产品", "高", "门票+接送"

    if has_ticket_word and re.search(r"subway|combo|optional|voucher|cable car|round-trip bus", t, re.I):
        return "门票产品", "高", "门票组合（门票+地铁/餐券/缆车等）"

    if has_ticket_word:
        return "门票产品", "高", "单门票"

    # 日游：group / bus / 途益。任务说日游目前主要是途益 bus tour
    if DAYTOUR_RE.search(t):
        return "日游产品", "高", "Group/Bus tour（日游；途益 bus tour 也归这里）"

    # 包车：Private Tour、按小时、指定车型、司机/导游
    if CHARTER_RE.search(t) or PRIVATE_TOUR_RE.search(t):
        return "包车产品", "中", "Private Tour / 按小时 / 车+司机，按任务先归包车（option 若是集合点再改日游）"

    # 10-Hour ... Tour 但没写 Private，更像团进团出日游
    if HOURLY_TOUR_RE.search(t):
        return "日游产品", "中", "按小时团游、未写 Private，先归日游"

    if SPA_OR_VENUE_RE.search(t) or re.search(r"\bpass\b", t, re.I):
        return "门票产品", "中", "场馆/温泉/通行证，先按门票"

    return "待确认", "低", "标题无法判断，需要看 option / 产品详情"


# merge page 0 if provided via stdin file
page0_path = DATA / "viator-api-page-0.json"
if page0_path.exists():
    raw = json.loads(page0_path.read_text())
    body = json.loads(raw["text"]) if isinstance(raw.get("text"), str) else raw
    for p in body.get("products") or []:
        by_code[p["productCode"]] = p

all_products = [slim(p) for p in by_code.values()]
active = [p for p in all_products if p["status"] == "ACTIVE" or p["isActive"] is True]
active.sort(key=lambda p: (p["productCode"] or ""), reverse=True)

rows = []
for p in active:
    cat, conf, reason = classify(p["title"])
    rows.append({**p, "category": cat, "confidence": conf, "reason": reason})

# persist
payload = {
    "generatedAt": datetime.now(timezone.utc).isoformat(),
    "allCount": len(all_products),
    "activeCount": len(active),
    "statusCounts": dict(Counter(p["status"] for p in all_products)),
    "categoryCounts": dict(Counter(r["category"] for r in rows)),
    "confidenceCounts": dict(Counter(r["confidence"] for r in rows)),
    "products": rows,
}
(DATA / "viator-active-classified.json").write_text(json.dumps(payload, ensure_ascii=False, indent=2))

# excel for user review
wb = openpyxl.Workbook()
# summary
ws = wb.active
ws.title = "分类总览"
headers = ["产品代码", "产品名称", "建议分类", "把握", "分类依据", "已连接option数", "启用option数", "Viator链接"]
fills = {
    "门票产品": PatternFill("solid", fgColor="FFF2CC"),
    "接送产品": PatternFill("solid", fgColor="D0E2FF"),
    "包车产品": PatternFill("solid", fgColor="E2D5F1"),
    "日游产品": PatternFill("solid", fgColor="D9EAD3"),
    "待确认": PatternFill("solid", fgColor="F4CCCC"),
}
conf_fill = {
    "高": PatternFill("solid", fgColor="D9EAD3"),
    "中": PatternFill("solid", fgColor="FFF2CC"),
    "低": PatternFill("solid", fgColor="F4CCCC"),
}
header_fill = PatternFill("solid", fgColor="1F4E79")
header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
body_font = Font(name="Calibri", size=11)
thin = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)

for col, h in enumerate(headers, 1):
    cell = ws.cell(1, col, h)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(vertical="center")

for i, r in enumerate(rows, 2):
    vals = [
        r["productCode"],
        r["title"],
        r["category"],
        r["confidence"],
        r["reason"],
        r["connectedOptionCount"],
        r["totalActiveOptionCount"],
        r["viatorUrl"],
    ]
    for c, v in enumerate(vals, 1):
        cell = ws.cell(i, c, v)
        cell.font = body_font
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = thin
        if c == 3:
            cell.fill = fills.get(r["category"], PatternFill())
        elif c == 4:
            cell.fill = conf_fill.get(r["confidence"], PatternFill())
        else:
            cell.fill = fills.get(r["category"], PatternFill())

ws.auto_filter.ref = f"A1:H{len(rows)+1}"
ws.freeze_panes = "A2"
widths = [16, 72, 14, 8, 42, 16, 14, 42]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.row_dimensions[1].height = 22
for i in range(2, len(rows) + 2):
    ws.row_dimensions[i].height = 28

# per-category sheets
order = ["门票产品", "接送产品", "包车产品", "日游产品", "待确认"]
for cat in order:
    wsc = wb.create_sheet(cat)
    for col, h in enumerate(["产品代码", "产品名称", "把握", "分类依据", "已连接option数", "Viator链接"], 1):
        cell = wsc.cell(1, col, h)
        cell.fill = header_fill
        cell.font = header_font
    cat_rows = [r for r in rows if r["category"] == cat]
    for i, r in enumerate(cat_rows, 2):
        vals = [r["productCode"], r["title"], r["confidence"], r["reason"], r["connectedOptionCount"], r["viatorUrl"]]
        for c, v in enumerate(vals, 1):
            cell = wsc.cell(i, c, v)
            cell.font = body_font
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = thin
            cell.fill = fills[cat]
    wsc.freeze_panes = "A2"
    wsc.auto_filter.ref = f"A1:F{max(1, len(cat_rows)+1)}"
    for i, w in enumerate([16, 72, 8, 48, 16, 42], 1):
        wsc.column_dimensions[get_column_letter(i)].width = w

# counts sheet
wss = wb.create_sheet("统计", 0)
wss["A1"] = "Viator 已启用产品分类（待你确认）"
wss["A1"].font = Font(name="Calibri", size=16, bold=True, color="1F4E79")
wss["A2"] = f"生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M')}   已启用 {len(active)} / 后台全部 {len(all_products)}"
wss["A3"] = "高=标题很明确；中=大方向对但可能边界情况；低=必须你拍板。低把握的请先看「待确认」和各表里标红的。"
wss.merge_cells("A1:C1")
wss.merge_cells("A2:C2")
wss.merge_cells("A3:C3")
wss["A5"] = "分类"
wss["B5"] = "数量"
wss["C5"] = "其中低把握"
for col in range(1, 4):
    wss.cell(5, col).fill = header_fill
    wss.cell(5, col).font = header_font
for i, cat in enumerate(order, 6):
    n = sum(1 for r in rows if r["category"] == cat)
    low = sum(1 for r in rows if r["category"] == cat and r["confidence"] == "低")
    wss.cell(i, 1, cat).fill = fills[cat]
    wss.cell(i, 2, n).fill = fills[cat]
    wss.cell(i, 3, low).fill = fills[cat]
    wss.cell(i, 1).font = body_font
    wss.cell(i, 2).font = body_font
    wss.cell(i, 3).font = body_font
wss.column_dimensions["A"].width = 16
wss.column_dimensions["B"].width = 10
wss.column_dimensions["C"].width = 14

out = ROOT / "Viator已启用产品分类_待确认.xlsx"
wb.save(out)
print("all", len(all_products), "active", len(active))
print("categories", payload["categoryCounts"])
print("confidence", payload["confidenceCounts"])
print("saved", out)
print("json", DATA / "viator-active-classified.json")
# print uncertain
print("\n=== 低把握 / 待确认 ===")
for r in rows:
    if r["category"] == "待确认" or r["confidence"] == "低":
        print(f"{r['confidence']:2} {r['category']:6} {r['productCode']} {r['title']} | {r['reason']}")