#!/usr/bin/env python3
"""Fill 门票表 with conservative Ctrip/Klook matches. Wrong-type hits are dropped."""
from __future__ import annotations

import json
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path("/Users/mac/viator产品爬取")
options = json.loads((ROOT / "data/viator-options.json").read_text())
matches = json.loads((ROOT / "data/ticket-supply-matches.json").read_text())
ck = json.loads((ROOT / "data/ticket-supply-checkpoint.json").read_text())
by_match = {r["productCode"]: r for r in matches["results"]}

CTRIP = "https://travelagents.trip.com/ttddist/act/dest/t{}.html"
KLOOK = "https://klook.klktech.com/activity/{}"

# Official / corrected links. Empty supplier if we cannot confidently match.
# metro official (task example)
METRO_CTRIP = {
    "supplier": "携程",
    "url": CTRIP.format(24465457),
    "name": "Tokyo Metro 24/48/72-hour travel pass",
    "usage": "电子票入场",
    "cancel": "不可取消",
}
METRO_KLOOK = {
    "supplier": "klook",
    "url": KLOOK.format(1552),
    "name": "东京地铁三日券／二日券／一日券",
    "usage": "电子票入场",
    "cancel": "条件取消",
}

def K(id_, name, usage="电子票入场", cancel="条件取消"):
    return {"supplier": "klook", "url": KLOOK.format(id_), "name": name, "usage": usage, "cancel": cancel}

def C(id_, name, usage="电子票入场", cancel="条件取消"):
    return {"supplier": "携程", "url": CTRIP.format(id_), "name": name, "usage": usage, "cancel": cancel}

# product-level park/attraction ticket (not subway, not dining)
PARK = {
    "5514894P9": [K(251, "首尔乐天世界门票"), C(110082115, "Sky Lotte World Tower - Ticket", cancel="不可取消")],  # combo handled in split
    "5514894P7": [K(39, "香港迪士尼乐园门票")],
    "5514894P61": [K(56183, "北京环球影城门票")],
    "5514894P60": [K(2128, "上海迪士尼乐园门票")],
    "5514894P58": [K(13283, "岘港太阳世界巴拿山门票")],
    "5514894P57": [K(27021, "胡志明市Landmark 81 Saigon Skyview门票"), C(111100087, "Landmark 81: SkyView Entry + VR Experience")],
    "5514894P55": [K(506, "澳门旅游塔观景门票"), C(87573427, "Macau Tower admission ticket", cancel="不可取消")],
    "5514894P54": [K(33776, "澳门 teamLab 超自然空间门票")],
    "5514894P53": [K(90, "澳门水舞间水上汇演门票")],
    "5514894P52": [K(695, "东京迪士尼度假区门票")],
    "5514894P51": [K(695, "东京迪士尼度假区门票")],
    "5514894P502": [K(149037, "teamLab Phenomena 阿布扎比门票")],
    "5514894P489": [K(4911, "日本东京铁塔瞭望台门票")],
    "5514894P483": [K(4911, "日本东京铁塔瞭望台门票")],
    "5514894P477": [K(695, "东京迪士尼度假区门票")],
    "5514894P475": [K(695, "东京迪士尼度假区门票")],
    "5514894P473": [K(13283, "岘港太阳世界巴拿山门票")],
    "5514894P47": [K(44688, "日本乐高度假乐园门票")],
    "5514894P468": [K(13283, "岘港太阳世界巴拿山门票")],
    "5514894P463": [K(84374, "东京哈利波特制片厂之旅")],
    "5514894P461": [K(84374, "东京哈利波特制片厂之旅")],
    "5514894P44": [K(695, "东京迪士尼度假区门票")],
    "5514894P437": [K(16738, "上海豫园门票")],
    "5514894P424": [K(695, "东京迪士尼度假区门票")],
    "5514894P423": [],  # no official Shibuya Sky ticket found
    "5514894P417": [],
    "5514894P404": [K(84374, "东京哈利波特制片厂之旅")],
    "5514894P403": [K(695, "东京迪士尼度假区门票")],
    "5514894P38": [K(123296, "泡泡玛特城市乐园")],
    "5514894P3": [K(695, "东京迪士尼度假区门票"), C(96477084, "Tokyo DisneyLand Park 1-Day Ticket", cancel="不可取消")],
    "5514894P299": [K(695, "东京迪士尼度假区门票"), C(96477084, "Tokyo DisneyLand Park 1-Day Ticket", cancel="不可取消")],
    "5514894P283": [K(39, "香港迪士尼乐园门票")],
    "5514894P282": [K(695, "东京迪士尼度假区门票")],
    "5514894P28": [K(1417, "有马温泉 太合之汤")],
    "5514894P251": [K(38494, "上海欢乐谷门票")],
    "5514894P247": [K(3972, "金茂大厦88层观光厅门票")],
    "5514894P245": [K(6365, "上海ERA时空之旅2杂技表演门票"), C(95752848, "Shanghai Circus World Tickets: ERA2")],
    "5514894P244": [K(4333, "上海中心大厦118楼观景台门票")],
    "5514894P24": [K(81280, "釜山海云台X the Sky门票")],
    "5514894P228": [K(39, "香港迪士尼乐园门票")],
    "5514894P21": [K(17678, "首尔乐天世界塔 Seoul Sky 观景台门票")],
    "5514894P184": [K(28956, "北京恭王府景区门票")],
    "5514894P183": [K(9808, "北京圆明园门票")],
    "5514894P18": [K(412, "N首尔塔门票")],
    "5514894P177": [K(149738, "北京老舍茶馆综艺演出票")],
    "5514894P163": [K(35861, "大阪梅田蓝天大厦 & 空中庭园观景台门票")],
    "5514894P162": [K(2424, "阿倍野展望台门票 HARUKAS 300")],
    "5514894P13": [K(695, "东京迪士尼度假区门票")],
    "5514894P115": [K(695, "东京迪士尼度假区门票")],
    "5514894P11": [K(695, "东京迪士尼度假区门票"), C(96477084, "Tokyo DisneyLand Park 1-Day Ticket", cancel="不可取消")],
}

MEAL_P7 = C(107942030, "Hong Kong Disneyland dining e-ticket", cancel="条件取消")
LOTTE_WORLD = K(251, "首尔乐天世界门票")
SEOUL_SKY = [C(110082115, "Sky Lotte World Tower - Ticket", cancel="不可取消"), K(17678, "首尔乐天世界塔 Seoul Sky 观景台门票")]


def splits_for(p, opt_title: str) -> list[tuple[str, list[dict]]]:
    """Return [(拆分, [supply dicts])] for one option."""
    t = f"{p['title']} {opt_title}"
    code = p["productCode"]
    park = list(PARK.get(code) or [])
    has_sub = bool(__import__("re").search(r"subway|metro|地铁", t, __import__("re").I))
    has_meal = bool(__import__("re").search(r"lunch|snack|voucher|餐|takoyaki", t, __import__("re").I))
    has_xfer = bool(__import__("re").search(r"transfer|pick-?up|drop-?off|shuttle|bus from|接送", t, __import__("re").I))
    ticket_only = bool(__import__("re").search(r"ticket only|no subway|admission only|ba na hills admission", opt_title, __import__("re").I))

    # P9 two attractions
    if code == "5514894P9":
        return [("乐天世界门票", [LOTTE_WORLD]), ("Seoul Sky门票", SEOUL_SKY)]

    out: list[tuple[str, list[dict]]] = []
    if park:
        out.append(("景区门票", park))
    else:
        out.append(("景区门票", []))

    if has_sub and not ticket_only:
        out.append(("地铁票", [METRO_CTRIP, METRO_KLOOK]))
    if has_meal and code == "5514894P7":
        out.append(("餐券", [MEAL_P7]))
    elif has_meal and code == "5514894P162":
        out.append(("餐食", []))  # takoyaki add-on, no standalone official SKU
    if has_xfer and not ticket_only:
        out.append(("接送", []))  # self-operated; do not match tour/transfer SKUs
    return out


YELLOW = PatternFill("solid", fgColor="FFF258")
HEADER = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
BODY = Font(name="Calibri", size=11)
LINK_FONT = Font(name="Calibri", size=11, color="0563C1", underline="single")
THIN = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)
WRAP = Alignment(vertical="center", wrap_text=True)


def paint(ws, r, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(r, c)
        cell.font = BODY if c != 8 else (LINK_FONT if cell.value else BODY)
        cell.alignment = WRAP
        cell.border = THIN
        cell.fill = YELLOW
    ws.row_dimensions[r].height = 24


wb = openpyxl.Workbook()

# summary
wss = wb.active
wss.title = "匹配说明"
wss["A1"] = "门票供应链匹配（携程 / Klook）"
wss["A1"].font = Font(name="Calibri", size=16, bold=True, color="1F4E79")
wss["A3"] = datetime.now().strftime("生成时间：%Y-%m-%d %H:%M")
notes = [
    "只保留能对上「官方门票」的采购链接；tour / 接送 / 巴士 / 船票 / 城市通票套餐已剔除。",
    "地铁票统一用示例同款：携程 Tokyo Metro 24/48/72-hour travel pass（24465457），并附 Klook 东京地铁券。",
    "Tokyo Tower 用 Klook 东京铁塔瞭望台（4911），不用晴空塔。",
    "Pop Land 用 Klook 泡泡玛特城市乐园（123296），不用环球影城。",
    "Shibuya Sky 两边都没有官方单票，景区门票链接留空，避免错配 MAGNET / 晴空塔 / 观光巴士。",
    "门票+接送：只配景区门票；接送视为自营，不配 tour/接送产品。",
    "Klook 使用方式按电子票入场；取消规则能写死的写死，否则写条件取消。",
]
for i, n in enumerate(notes, 5):
    wss.cell(i, 1, n)
    wss.merge_cells(start_row=i, start_column=1, end_row=i, end_column=8)
wss.column_dimensions["A"].width = 110

# tickets sheet
ws = wb.create_sheet("门票产品")
headers = ["产品名称", "产品代码", "option", "具体拆分", "门票使用方式", "门票取消规则", "供应商", "供应商链接"]
for col, h in enumerate(headers, 1):
    cell = ws.cell(1, col, h)
    cell.fill = HEADER
    cell.font = HEADER_FONT
    cell.alignment = Alignment(vertical="center")
ws.freeze_panes = "A2"
ws.row_dimensions[1].height = 22

tickets = [p for p in options["products"] if p["category"] == "门票产品"]
r = 2
unmatched = []
for p in tickets:
    opts = [o for o in (p.get("options") or []) if (o.get("status") or "ACTIVE") == "ACTIVE"] or [{}]
    prod_start = r
    for opt in opts:
        opt_title = (opt.get("title") or "").strip() or "DEFAULT"
        groups = splits_for(p, opt_title)
        opt_start = r
        for split_name, supplies in groups:
            if not supplies:
                ws.cell(r, 1, p["title"])
                ws.cell(r, 2, p["productCode"])
                ws.cell(r, 3, opt_title)
                ws.cell(r, 4, split_name)
                paint(ws, r, 8)
                if split_name == "景区门票":
                    unmatched.append((p["productCode"], p["title"], opt_title))
                r += 1
                continue
            for s in supplies:
                ws.cell(r, 1, p["title"])
                ws.cell(r, 2, p["productCode"])
                ws.cell(r, 3, opt_title)
                ws.cell(r, 4, split_name)
                ws.cell(r, 5, s.get("usage") or "")
                ws.cell(r, 6, s.get("cancel") or "")
                ws.cell(r, 7, s.get("supplier") or "")
                ws.cell(r, 8, s.get("url") or "")
                paint(ws, r, 8)
                r += 1
        if r - 1 > opt_start:
            ws.merge_cells(start_row=opt_start, start_column=3, end_row=r - 1, end_column=3)
            ws.cell(opt_start, 3).alignment = Alignment(vertical="center", wrap_text=True)
    if r - 1 > prod_start:
        ws.merge_cells(start_row=prod_start, start_column=1, end_row=r - 1, end_column=1)
        ws.merge_cells(start_row=prod_start, start_column=2, end_row=r - 1, end_column=2)
        ws.cell(prod_start, 1).alignment = Alignment(vertical="center", wrap_text=True)
        ws.cell(prod_start, 2).alignment = Alignment(vertical="center", wrap_text=True)

ws.auto_filter.ref = f"A1:H{max(2, r - 1)}"
for i, w in enumerate([54, 16, 36, 14, 14, 16, 10, 52], 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# copy other 3 sheets from option workbook
opt_wb = openpyxl.load_workbook(ROOT / "Viator已启用产品_option.xlsx")
for name in ["接送产品", "包车产品", "日游产品"]:
    src = opt_wb[name]
    dst = wb.create_sheet(name)
    for row in src.iter_rows():
        for cell in row:
            dst.cell(cell.row, cell.column, cell.value)
            if cell.has_style:
                dst.cell(cell.row, cell.column).font = cell.font.copy()
                dst.cell(cell.row, cell.column).fill = cell.fill.copy()
                dst.cell(cell.row, cell.column).alignment = cell.alignment.copy()
                dst.cell(cell.row, cell.column).border = cell.border.copy()
    for m in src.merged_cells.ranges:
        dst.merge_cells(str(m))
    for col, dim in src.column_dimensions.items():
        dst.column_dimensions[col].width = dim.width
    dst.freeze_panes = src.freeze_panes

out = ROOT / "Viator已启用产品_门票已匹配.xlsx"
wb.save(out)
print("saved", out)
print("ticket rows", r - 2)
print("unmatched 景区门票", len(unmatched))
for u in unmatched:
    print(" ", u)
